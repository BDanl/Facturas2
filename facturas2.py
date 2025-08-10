import sys
import json
import csv
import locale
import logging
import os
import ctypes
from ctypes import wintypes
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict
import webbrowser
import configparser
from database import Database
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo

# Importaciones de PyQt6
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QPushButton, QComboBox, QDateEdit, QAbstractItemView, 
                             QTableWidget, QTableWidgetItem, QTabWidget, QMessageBox, 
                             QFileDialog, QHeaderView, QTextEdit, QCheckBox, QSplitter,
                             QStyleFactory, QStyle, QTableWidgetSelectionRange, QStatusBar,
                             QGroupBox, QFormLayout, QSpacerItem, QSizePolicy, QTreeWidget, 
                             QTreeWidgetItem, QMenu, QDialog, QListWidget, QDialogButtonBox, 
                             QListWidgetItem, QProgressDialog, QStyledItemDelegate)
from PyQt6.QtGui import (QAction, QFont, QColor, QIcon, QDoubleValidator, 
                        QTextCursor, QBrush)
from PyQt6.QtCore import Qt, QSize, QDate, QTimer, QModelIndex


class EditableDelegate(QStyledItemDelegate):
    """Delegate para controlar qué celdas son editables"""
    def __init__(self, parent=None, editable_columns=None):
        super().__init__(parent)
        self.editable_columns = editable_columns or []
    
    def createEditor(self, parent, option, index):
        """Sobrescribir para crear el editor solo para columnas editables"""
        if index.column() in self.editable_columns:
            editor = super().createEditor(parent, option, index)
            # Aumentar el tamaño mínimo del editor
            if isinstance(editor, QLineEdit):
                editor.setMinimumHeight(40)  # Altura de 40 píxeles
                editor.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.MinimumExpanding)
            return editor
        return None
    
    def updateEditorGeometry(self, editor, option, index):
        """Actualizar la geometría del editor para que ocupe más espacio"""
        if index.column() in self.editable_columns:
            # Aumentar la altura del área de edición
            option.rect.setHeight(max(40, option.rect.height()))
            editor.setGeometry(option.rect)
    
    def setEditorData(self, editor, index):
        """Cargar datos en el editor"""
        if index.column() in self.editable_columns:
            super().setEditorData(editor, index)
    
    def setModelData(self, editor, model, index):
        """Guardar datos del editor en el modelo"""
        if index.column() in self.editable_columns:
            super().setModelData(editor, model, index)


class TipoGastoDelegate(QStyledItemDelegate):
    """Delegate para la columna de tipo de gasto con QComboBox"""
    def __init__(self, parent=None, tipos_gasto=None, column_index=2):
        super().__init__(parent)
        self.tipos_gasto = tipos_gasto or []
        self.column_index = column_index  # Índice de la columna de tipo
    
    def createEditor(self, parent, option, index):
        """Crear un QComboBox como editor"""
        if index.column() == self.column_index:  # Usar el índice de columna configurado
            editor = QComboBox(parent)
            # Agregar los tipos de gasto al combobox
            for tipo in self.tipos_gasto:
                editor.addItem(tipo['nombre'])
            editor.setMinimumHeight(40)
            editor.installEventFilter(self)
            return editor
        return super().createEditor(parent, option, index)
    
    def setEditorData(self, editor, index):
        """Establecer el valor actual en el editor"""
        if index.column() == self.column_index:
            current_text = index.data(Qt.ItemDataRole.DisplayRole)
            idx = editor.findText(current_text)
            if idx >= 0:
                editor.setCurrentIndex(idx)
        else:
            super().setEditorData(editor, index)
    
    def setModelData(self, editor, model, index):
        """Guardar el valor seleccionado en el modelo"""
        if index.column() == self.column_index:
            model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)
        else:
            super().setModelData(editor, model, index)
    
    def updateEditorGeometry(self, editor, option, index):
        """Ajustar la geometría del editor"""
        if index.column() == self.column_index:
            editor.setGeometry(option.rect)
        else:
            super().updateEditorGeometry(editor, option, index)



# Importaciones para Excel
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as e:
    logging.error(f"Error al importar dependencias de Excel: {str(e)}")
    QMessageBox.critical(None, "Error", "Error al importar dependencias de Excel. Asegúrate de tener openpyxl instalado.")
    sys.exit(1)

    # Crear directorio de datos si no existe
    data_dir = Path.home() / 'FacturasApp'
    data_dir.mkdir(exist_ok=True, parents=True)
    
def check_single_instance():
    """Verifica si ya hay una instancia de la aplicación en ejecución"""
    # Usar un nombre único para el mutex
    mutex_name = f"Global\\{APP_NAME}_SingleInstanceMutex"
    
    # Crear un mutex con nombre
    mutex = ctypes.windll.kernel32.CreateMutexW(None, False, mutex_name)
    
    # Verificar si el mutex ya existía
    last_error = ctypes.windll.kernel32.GetLastError()
    
    if last_error == 183:  # ERROR_ALREADY_EXISTS
        # Obtener el handle de la ventana existente
        hwnd = ctypes.windll.user32.FindWindowW(None, "Gestor de Facturas")
        if hwnd:
            # Si la ventana está minimizada, restaurarla
            if ctypes.windll.user32.IsIconic(hwnd):
                ctypes.windll.user32.ShowWindow(hwnd, 9)  # SW_RESTORE = 9
            
            # Traer al frente
            ctypes.windll.user32.SetForegroundWindow(hwnd)
            
            # Activar la ventana
            ctypes.windll.user32.SetActiveWindow(hwnd)
            
            # Traer al frente de las demás ventanas
            ctypes.windll.user32.SetWindowPos(
                hwnd,  # hWnd
                -1,    # hWndInsertAfter (HWND_TOPMOST)
                0, 0, 0, 0,  # x, y, cx, cy (ignorados)
                0x0001 | 0x0002  # SWP_NOMOVE | SWP_NOSIZE
            )
        return False
    
    # Configurar para que el mutex no se cierre automáticamente al salir
    ctypes.windll.kernel32.SetHandleInformation(mutex, 1, 1)
    return True

def main():
    # Verificar si ya hay una instancia en ejecución
    mutex = ctypes.windll.kernel32.CreateMutexW(None, False, "Global\\GestorFacturas_SingleInstance")
    
    # Verificar si el mutex ya existía
    error = ctypes.windll.kernel32.GetLastError()
    
    if error == 183:  # ERROR_ALREADY_EXISTS
        # Encontrar la ventana existente
        hwnd = ctypes.windll.user32.FindWindowW(None, "Gestor de Facturas")
        if hwnd:
            # Restaurar si está minimizada
            if ctypes.windll.user32.IsIconic(hwnd):
                ctypes.windll.user32.ShowWindow(hwnd, 9)  # SW_RESTORE = 9
            
            # Traer al frente y activar
            ctypes.windll.user32.SetForegroundWindow(hwnd)
            ctypes.windll.user32.SetActiveWindow(hwnd)
            
            # Forzar el enfoque
            ctypes.windll.user32.BringWindowToTop(hwnd)
            ctypes.windll.user32.SetFocus(hwnd)
            
            # Mensaje de depuración
            logger.warning("Se intentó abrir una segunda instancia. Redirigiendo a la ventana existente.")
        
        # Salir de la nueva instancia
        sys.exit(0)
    
# Configuración de directorios de la aplicación
APP_NAME = "GestorFacturas"
DATA_DIR = Path.home() / 'FacturasApp'
DATA_DIR.mkdir(exist_ok=True, parents=True)

# Configurar logging a nivel de módulo
log_file = DATA_DIR / 'gestor_facturas.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(APP_NAME)

# Configuración de la aplicación
CONFIG_FILE = str(DATA_DIR / 'config.ini')

def get_config():
    """Obtener la configuración de la aplicación"""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
    
    # Configuración por defecto
    if 'APP' not in config:
        config['APP'] = {}
    if 'last_export_dir' not in config['APP']:
        config['APP']['last_export_dir'] = str(Path.home() / 'Documents')
    
    return config

def save_config(config):
    """Guardar la configuración de la aplicación"""
    with open(CONFIG_FILE, 'w') as configfile:
        config.write(configfile)

# Configurar formato de moneda colombiana
try:
    locale.setlocale(locale.LC_ALL, 'es_CO.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Spanish_Colombia.1252')
    except:
        locale.setlocale(locale.LC_ALL, '')

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestor de Facturas")
        # Remove fixed geometry and use showMaximized() after UI is initialized
        
        # Determinar si estamos ejecutando desde un ejecutable o desde el código fuente
        if getattr(sys, 'frozen', False):
            # Si es un ejecutable, usamos el directorio del ejecutable
            self.app_dir = Path(sys._MEIPASS) if hasattr(sys, '_MEIPASS') else Path(sys.executable).parent
            self.data_dir = Path.home() / "FacturasApp"
        else:
            # Si es código fuente, usamos el directorio del proyecto
            self.app_dir = Path(__file__).parent
            self.data_dir = self.app_dir
        
        # Asegurarse de que el directorio de datos exista
        self.data_dir.mkdir(exist_ok=True, parents=True)
        
        # Inicializar la base de datos SQLite en el directorio de datos
        db_path = DATA_DIR / "facturas.db"
        self.db = Database(str(db_path))
        
        # Verificar si hay que migrar datos desde el archivo JSON antiguo
        self._migrar_datos_desde_json()
        
        # Configuración del tema
        self.tema_oscuro = False
        
        # Inicializar atributos
        self.facturas = []
        self.tipos_gasto = []
        
        # Cargar datos sin actualizar la UI aún
        self.cargar_datos(actualizar_ui=False)
        
        # Inicializar la interfaz de usuario
        self.init_ui()
        
        # Aplicar el tema antes de mostrar la ventana
        self.aplicar_tema()
        
        # Actualizar la UI con los datos cargados
        self.actualizar_lista_facturas()
        self.actualizar_filtros()
        self.actualizar_resumen()
        
        # Maximizar la ventana después de inicializar la UI
        self.showMaximized()
    
    def _migrar_datos_desde_json(self):
        """Migra los datos desde el archivo JSON antiguo a la base de datos SQLite si es necesario."""
        json_path = Path.home() / "FacturasApp" / "facturas_qt.json"
        if json_path.exists():
            try:
                # Verificar si ya hay datos en la base de datos
                facturas = self.db.obtener_facturas()
                if not facturas:
                    # Migrar solo si no hay datos en la base de datos
                    self.db.migrar_desde_json(str(json_path))
                    # Opcional: respaldar el archivo JSON después de la migración
                    backup_path = json_path.with_suffix(f".{datetime.now().strftime('%Y%m%d%H%M%S')}.json")
                    json_path.rename(backup_path)
                    logging.info(f"Datos migrados exitosamente. Archivo original respaldado como {backup_path}")
            except Exception as e:
                logging.error(f"Error al migrar datos desde JSON: {str(e)}")
    
    def init_ui(self):
        """Inicializar la interfaz de usuario"""
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(10)
        
        # Barra de herramientas
        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(0, 0, 0, 10)
        
        # Título
        title = QLabel("Sistema de Gestión de Facturas")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Botón para alternar modo oscuro/claro
        self.btn_tema = QPushButton("Modo Oscuro")
        self.btn_tema.setToolTip("Haz clic para cambiar entre modo claro y oscuro")
        self.btn_tema.setFixedSize(120, 32)
        self.btn_tema.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 5px 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
            QPushButton:pressed {
                background-color: #4e555b;
            }
        """)
        self.btn_tema.clicked.connect(self.cambiar_tema)
        
        # Agregar widgets a la barra de herramientas
        toolbar.addStretch()
        toolbar.addWidget(title)
        toolbar.addStretch()
        toolbar.addWidget(self.btn_tema)
        
        # Agregar barra de herramientas al layout principal
        main_layout.addLayout(toolbar)
        
        # Inicializar tema (por defecto: modo claro)
        self.cargar_preferencia_tema()
        
        # Configurar pestañas
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # Pestaña de registro
        self.tab_registro = QWidget()
        self.tabs.addTab(self.tab_registro, "Registrar Factura")
        self.setup_registro_tab()
        
        # Pestaña de resúmenes (antes Filtros Avanzados)
        self.tab_filtros = QWidget()
        self.tabs.addTab(self.tab_filtros, "Resúmenes")
        self.setup_filtros_tab()
        
        # Pestaña de lista de facturas
        self.tab_lista = QWidget()
        self.tabs.addTab(self.tab_lista, "Lista de Facturas")
        self.setup_lista_tab()
        
        # Barra de estado
        self.statusBar().showMessage("Listo")
    
    def setup_registro_tab(self):
        """Configurar la pestaña de registro de facturas"""
        layout = QFormLayout(self.tab_registro)
        
        # Tipo de gasto
        self.cmb_tipo_gasto = QComboBox()
        # Cargar tipos de gasto desde la base de datos
        try:
            tipos_gasto = self.db.obtener_tipos_gasto()
            for tipo in tipos_gasto:
                self.cmb_tipo_gasto.addItem(tipo['nombre'])
        except Exception as e:
            # En caso de error, cargar valores por defecto
            logging.error(f"Error al cargar tipos de gasto: {str(e)}")
            self.cmb_tipo_gasto.addItems(["Mercado", "Transporte", "Compra tienda", 
                                        "Farmacia", "Varios", "Gastos urgentes"])
        
        # Descripción
        self.txt_descripcion = QLineEdit()
        
        # Valor
        self.txt_valor = QLineEdit()
        self.txt_valor.setValidator(QDoubleValidator(0, 999999999, 2, self))
        
        # Fecha
        self.date_fecha = QDateEdit()
        self.date_fecha.setCalendarPopup(True)
        self.date_fecha.setDate(QDate.currentDate())
        self.date_fecha.setDisplayFormat("dd/MM/yyyy")
        
        # Botón de guardar
        self.btn_guardar = QPushButton("Guardar Factura")
        self.btn_guardar.clicked.connect(self.guardar_factura)
        
        # Agregar widgets al layout
        layout.addRow("Tipo de Gasto:", self.cmb_tipo_gasto)
        layout.addRow("Descripción:", self.txt_descripcion)
        layout.addRow("Valor (COP):", self.txt_valor)
        layout.addRow("Fecha:", self.date_fecha)
        layout.addRow(self.btn_guardar)
    
    def setup_resumen_tab(self):
        """Configurar la pestaña de resumen"""
        # Crear pestañas para diferentes vistas de resumen
        self.tabs_resumen = QTabWidget()
        
        # Pestaña de resumen diario
        self.tab_diario = QWidget()
        self.tabs_resumen.addTab(self.tab_diario, "Resumen Diario")
        self.setup_resumen_diario_tab()
        
        # Pestaña de resumen mensual
        self.tab_mensual = QWidget()
        self.tabs_resumen.addTab(self.tab_mensual, "Resumen Mensual")
        self.setup_resumen_mensual_tab()
        
        # Pestaña de resumen anual
        self.tab_anual = QWidget()
        self.tabs_resumen.addTab(self.tab_anual, "Resumen Anual")
        self.setup_resumen_anual_tab()
        
        # Pestaña de filtros avanzados
        self.tab_filtros = QWidget()
        self.tabs_resumen.addTab(self.tab_filtros, "Filtros Avanzados")
        self.setup_filtros_tab()
        
        # Layout principal
        layout = QVBoxLayout(self.tab_resumen)
        layout.addWidget(self.tabs_resumen)
        
        # Actualizar los resúmenes
        self.actualizar_resumen()
    
    def setup_resumen_diario_tab(self):
        """Configurar la pestaña de resumen diario"""
        layout = QVBoxLayout(self.tab_diario)
        
        # Fecha para el resumen diario
        self.date_resumen_diario = QDateEdit()
        self.date_resumen_diario.setCalendarPopup(True)
        self.date_resumen_diario.setDate(QDate.currentDate())
        self.date_resumen_diario.setDisplayFormat("dd/MM/yyyy")
        self.date_resumen_diario.dateChanged.connect(self.actualizar_resumen_diario)
        
        # Botón para ir a hoy
        btn_hoy = QPushButton("Hoy")
        btn_hoy.clicked.connect(lambda: self.date_resumen_diario.setDate(QDate.currentDate()))
        
        # Layout para controles de fecha
        fecha_layout = QHBoxLayout()
        fecha_layout.addWidget(QLabel("Seleccione una fecha:"))
        fecha_layout.addWidget(self.date_resumen_diario)
        fecha_layout.addWidget(btn_hoy)
        fecha_layout.addStretch()
        
        # Área de texto para el resumen
        self.texto_resumen_diario = QTextEdit()
        self.texto_resumen_diario.setReadOnly(True)
        
        # Agregar al layout principal
        layout.addLayout(fecha_layout)
        layout.addWidget(self.texto_resumen_diario)
    
    def setup_resumen_mensual_tab(self):
        """Configurar la pestaña de resumen mensual"""
        layout = QVBoxLayout(self.tab_mensual)
        
        # Controles para seleccionar mes y año
        control_layout = QHBoxLayout()
        
        # Combo para seleccionar mes
        self.combo_mes_resumen = QComboBox()
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        self.combo_mes_resumen.addItems(meses)
        self.combo_mes_resumen.setCurrentIndex(QDate.currentDate().month() - 1)
        self.combo_mes_resumen.currentIndexChanged.connect(self.actualizar_resumen_mensual)
        
        # Combo para seleccionar año
        self.combo_anio_resumen = QComboBox()
        anio_actual = QDate.currentDate().year()
        self.combo_anio_resumen.addItems([str(anio) for anio in range(anio_actual - 5, anio_actual + 1)])
        self.combo_anio_resumen.setCurrentText(str(anio_actual))
        self.combo_anio_resumen.currentTextChanged.connect(self.actualizar_resumen_mensual)
        
        # Agregar controles al layout
        control_layout.addWidget(QLabel("Mes:"))
        control_layout.addWidget(self.combo_mes_resumen)
        control_layout.addWidget(QLabel("Año:"))
        control_layout.addWidget(self.combo_anio_resumen)
        control_layout.addStretch()
        
        # Área de texto para el resumen
        self.texto_resumen_mensual = QTextEdit()
        self.texto_resumen_mensual.setReadOnly(True)
        
        # Agregar al layout principal
        layout.addLayout(control_layout)
        layout.addWidget(self.texto_resumen_mensual)
        
        # Cargar datos iniciales
        self.actualizar_resumen_mensual()
    
    def setup_resumen_anual_tab(self):
        """Configurar la pestaña de resumen anual"""
        layout = QVBoxLayout(self.tab_anual)
        
        # Controles para seleccionar año
        control_layout = QHBoxLayout()
        
        # Combo para seleccionar año
        self.combo_anio_anual = QComboBox()
        anio_actual = QDate.currentDate().year()
        self.combo_anio_anual.addItems([str(anio) for anio in range(anio_actual - 10, anio_actual + 1)])
        self.combo_anio_anual.setCurrentText(str(anio_actual))
        self.combo_anio_anual.currentTextChanged.connect(self.actualizar_resumen_anual)
        
        # Agregar controles al layout
        control_layout.addWidget(QLabel("Año:"))
        control_layout.addWidget(self.combo_anio_anual)
        control_layout.addStretch()
        
        # Área de texto para el resumen
        self.texto_resumen_anual = QTextEdit()
        self.texto_resumen_anual.setReadOnly(True)
        
        # Agregar al layout principal
        layout.addLayout(control_layout)
        layout.addWidget(self.texto_resumen_anual)
        
        # Cargar datos iniciales
        self.actualizar_resumen_anual()
    
    def setup_filtros_tab(self):
        """Configurar la pestaña de filtros avanzados"""
        layout = QVBoxLayout(self.tab_filtros)
        
        # Grupo para filtros
        grupo_filtros = QGroupBox("Filtrar por")
        form_filtros = QFormLayout()
        
        # Filtro por rango de fechas
        self.date_edit_desde = QDateEdit()
        self.date_edit_desde.setCalendarPopup(True)
        self.date_edit_desde.setDisplayFormat("dd/MM/yyyy")
        # Establecer fecha inicial como primer día del mes actual
        today = QDate.currentDate()
        first_day_of_month = QDate(today.year(), today.month(), 1)
        self.date_edit_desde.setDate(first_day_of_month)
        self.date_edit_desde.dateChanged.connect(self.aplicar_filtros)
        
        self.date_edit_hasta = QDateEdit()
        self.date_edit_hasta.setCalendarPopup(True)
        self.date_edit_hasta.setDisplayFormat("dd/MM/yyyy")
        # Establecer fecha final como hoy
        self.date_edit_hasta.setDate(today)
        self.date_edit_hasta.dateChanged.connect(self.aplicar_filtros)
        
        # Filtro por año
        self.combo_filtro_anio = QComboBox()
        self.combo_filtro_anio.addItem("Todos los años", None)
        self.combo_filtro_anio.currentIndexChanged.connect(self.aplicar_filtros)
        
        # Filtro por mes
        self.combo_filtro_mes = QComboBox()
        self.combo_filtro_mes.addItem("Todos los meses", None)
        self.combo_filtro_mes.addItems([f"{i:02d} - {mes}" for i, mes in enumerate([
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ], 1)])
        self.combo_filtro_mes.currentIndexChanged.connect(self.aplicar_filtros)
        
        # Filtro por día
        self.combo_filtro_dia = QComboBox()
        self.combo_filtro_dia.addItem("Todos los días", None)
        self.combo_filtro_dia.addItems([f"{i:02d}" for i in range(1, 32)])
        self.combo_filtro_dia.currentIndexChanged.connect(self.aplicar_filtros)
        
        # Filtro por tipo de gasto
        self.combo_filtro_tipo = QComboBox()
        self.combo_filtro_tipo.addItem("Todos los tipos", None)
        try:
            tipos_gasto = self.db.obtener_tipos_gasto()
            for tipo in tipos_gasto:
                self.combo_filtro_tipo.addItem(tipo['nombre'])
        except Exception as e:
            # En caso de error, cargar valores por defecto
            logging.error(f"Error al cargar tipos de gasto en filtros: {str(e)}")
            self.combo_filtro_tipo.addItems(["Mercado", "Transporte", "Compra tienda", 
                                          "Farmacia", "Varios", "Gastos urgentes"])
        self.combo_filtro_tipo.currentIndexChanged.connect(self.aplicar_filtros)
        
        # Botón para limpiar filtros
        btn_limpiar = QPushButton("Limpiar Filtros")
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        
        # Agregar controles al formulario
        form_filtros.addRow("Fecha desde:", self.date_edit_desde)
        form_filtros.addRow("Fecha hasta:", self.date_edit_hasta)
        form_filtros.addRow("Año:", self.combo_filtro_anio)
        form_filtros.addRow("Mes:", self.combo_filtro_mes)
        form_filtros.addRow("Día:", self.combo_filtro_dia)
        form_filtros.addRow("Tipo de gasto:", self.combo_filtro_tipo)
        form_filtros.addRow(btn_limpiar)
        
        grupo_filtros.setLayout(form_filtros)
        
        # Tabla para mostrar resultados filtrados
        self.tabla_filtrada = QTableWidget()
        self.tabla_filtrada.setColumnCount(4)
        self.tabla_filtrada.setHorizontalHeaderLabels(["Fecha", "Tipo", "Descripción", "Valor"])
        
        # Configurar el ancho de las columnas
        header = self.tabla_filtrada.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Fecha
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Tipo
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Descripción
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Valor
        
        # Configurar edición de celdas
        self.tabla_filtrada.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed)
        
        # Configurar delegado para columnas editables (excepto tipo)
        delegate = EditableDelegate(self.tabla_filtrada, editable_columns=[0, 2, 3])  # Fecha, Descripción, Valor
        
        # Configurar delegado personalizado para la columna de tipo (índice 1)
        self.tipo_delegate_filtros = TipoGastoDelegate(
            self.tabla_filtrada, 
            tipos_gasto=self.tipos_gasto,
            column_index=1  # Índice de la columna 'tipo' en la tabla filtrada
        )
        self.tabla_filtrada.setItemDelegateForColumn(1, self.tipo_delegate_filtros)
        
        # Asignar el delegado general para las demás columnas
        self.tabla_filtrada.setItemDelegate(delegate)
        
        # Conectar la señal itemChanged al manejador de cambios
        self.tabla_filtrada.itemChanged.connect(self.guardar_cambios_celda)
        
        # Agregar widgets al layout principal
        layout.addWidget(grupo_filtros)
        layout.addWidget(self.tabla_filtrada)
        
        # Inicializar filtros y aplicar automáticamente
        self.inicializar_filtros()
        self.aplicar_filtros()
    
    def setup_lista_tab(self):
        """Configurar la pestaña de lista de facturas"""
        
        layout = QVBoxLayout(self.tab_lista)
        
        # Layout para los botones de acción
        btn_layout = QHBoxLayout()
        
        # Menú desplegable para importar desde diferentes formatos
        self.menu_importar = QPushButton("Importar Facturas ▼")
        self.menu_importar.setObjectName("menu_importar")
        self.menu_importar.setToolTip("Importar facturas desde diferentes formatos")
        self.menu_importar.setMinimumWidth(400)  # Ancho mínimo de 350 píxeles
        self.menu_importar.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
        self.menu_importar.setStyleSheet("""
            QPushButton {
                padding: 8px 15px;
                margin: 2px;
                min-width: 400px;
                font-size: 12px;
            }
        """)
        
        # Crear menú desplegable
        self.import_menu = QMenu(self)
        
        # Acciones del menú
        accion_json = QAction("Importar desde JSON", self)
        accion_csv = QAction("Importar desde CSV", self)
        accion_excel = QAction("Importar desde Excel", self)
        
        # Conectar acciones
        accion_json.triggered.connect(self.importar_desde_json)
        accion_csv.triggered.connect(self.importar_desde_csv)
        accion_excel.triggered.connect(self.importar_desde_excel)
        
        # Agregar acciones al menú
        self.import_menu.addAction(accion_json)
        self.import_menu.addAction(accion_csv)
        self.import_menu.addAction(accion_excel)
        
        # Configurar el botón para mostrar el menú
        self.menu_importar.setMenu(self.import_menu)
        btn_layout.addWidget(self.menu_importar)
        
        # Botón para limpiar todo
        self.btn_limpiar_todo = QPushButton("Limpiar Todo")
        self.btn_limpiar_todo.setObjectName("btn_limpiar_todo")
        self.btn_limpiar_todo.clicked.connect(self.confirmar_limpiar_todo)
        self.btn_limpiar_todo.setToolTip("Eliminar todas las facturas")
        self.btn_limpiar_todo.setMinimumWidth(400)  # Ancho mínimo de 350 píxeles
        self.btn_limpiar_todo.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
        self.btn_limpiar_todo.setStyleSheet("""
            QPushButton {
                padding: 8px 15px;
                margin: 2px;
                min-width: 400px;
                font-size: 12px;
            }
        """)
        btn_layout.addWidget(self.btn_limpiar_todo)
        
        # Botón para exportar a Excel
        self.btn_exportar = QPushButton("Exportar a Excel")
        self.btn_exportar.setObjectName("btn_exportar")
        self.btn_exportar.clicked.connect(self.exportar_a_excel)
        self.btn_exportar.setToolTip("Exportar facturas a un archivo Excel")
        self.btn_exportar.setMinimumWidth(400)  # Ancho mínimo de 350 píxeles
        self.btn_exportar.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
        self.btn_exportar.setStyleSheet("""
            QPushButton {
                padding: 8px 15px;
                margin: 2px;
                min-width: 400px;
                font-size: 12px;
            }
        """)
        btn_layout.addWidget(self.btn_exportar)
        
        # Layout para la tabla y botón de eliminar seleccionadas
        table_layout = QVBoxLayout()
        
        # Tabla de facturas
        self.tabla_facturas = QTableWidget()
        self.tabla_facturas.setColumnCount(5)  # Una columna extra para el ID
        self.tabla_facturas.setHorizontalHeaderLabels(["ID", "Fecha", "Tipo", "Descripción", "Valor"])
        self.tabla_facturas.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabla_facturas.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        
        # Conectar la señal de cambio de selección
        self.tabla_facturas.itemSelectionChanged.connect(self.actualizar_boton_eliminar)
        
        # Configurar las columnas
        self.tabla_facturas.setColumnHidden(0, True)  # Ocultar columna de checkboxes
        
        # Configurar delegados para celdas editables
        # Usar EditableDelegate para todas las columnas editables excepto tipo
        self.tabla_facturas.setItemDelegate(EditableDelegate(self.tabla_facturas, [1, 3, 4]))  # Columnas editables: Fecha (1), Descripción (3), Valor (4)
        
        # Configurar delegado personalizado para la columna de tipo (2)
        self.tabla_facturas.setItemDelegateForColumn(2, TipoGastoDelegate(self.tabla_facturas, self.tipos_gasto))
        
        # Ajustar el tamaño de las columnas
        header = self.tabla_facturas.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Columna oculta para checkboxes
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Fecha
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Tipo
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Descripción
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Valor
        
        # Configurar edición de celdas
        self.tabla_facturas.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed)
        
        # Conectar señal de cambio de celda
        self.tabla_facturas.itemChanged.connect(self.guardar_cambios_celda)
        
        # Layout para el botón de eliminar seleccionadas
        bottom_btn_layout = QHBoxLayout()
        self.btn_eliminar = QPushButton("Eliminar Seleccionadas")
        self.btn_eliminar.clicked.connect(self.eliminar_facturas_seleccionadas)
        self.btn_eliminar.setToolTip("Eliminar las facturas seleccionadas")
        self.btn_eliminar.setEnabled(False)  # Deshabilitado inicialmente
        self.btn_eliminar.setStyleSheet("""
            QPushButton {
                background-color: #f8d7da;
                color: #721c24;
                padding: 5px 10px;
                border: 1px solid #f5c6cb;
                border-radius: 4px;
            }
            QPushButton:disabled {
                background-color: #e2e3e5;
                color: #383d41;
                border: 1px solid #d6d8db;
            }
            QPushButton:hover:!disabled {
                background-color: #f5c6cb;
            }
        """)
        
        # Conectar señal de selección para habilitar/deshabilitar el botón
        self.tabla_facturas.itemSelectionChanged.connect(self.actualizar_boton_eliminar)
        
        bottom_btn_layout.addWidget(self.btn_eliminar)
        bottom_btn_layout.addStretch()
        
        # Agregar widgets al layout principal
        layout.addLayout(btn_layout)
        layout.addWidget(self.tabla_facturas)
        layout.addLayout(bottom_btn_layout)
    
    def cargar_datos(self):
        """Cargar datos desde la base de datos"""
        try:
            self.facturas = self.db.obtener_facturas()
            logger.info(f"Datos cargados correctamente desde la base de datos")
        except Exception as e:
            logger.error(f"Error al cargar los datos: {str(e)}")
            QMessageBox.critical(self, "Error", f"No se pudieron cargar los datos: {str(e)}")
    
    def guardar_datos(self):
        """Guardar datos en la base de datos"""
        try:
            self.db.guardar_facturas(self.facturas)
            return True
        except Exception as e:
            logger.error(f"Error al guardar datos: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Error al guardar datos: {str(e)}")
            return False
    
    def guardar_factura(self):
        """Guardar una nueva factura"""
        # Validar campos
        if not self.validar_campos():
            return
        
        # Crear diccionario con los datos de la factura
        factura = {
            'fecha': self.date_fecha.date().toString("dd/MM/yyyy"),
            'tipo': self.cmb_tipo_gasto.currentText(),
            'descripcion': self.txt_descripcion.text().strip(),
            'valor': float(self.txt_valor.text().replace(',', '.'))
        }
        
        # Agregar a la lista
        self.facturas.append(factura)
        
        # Guardar datos
        if self.guardar_datos():
            # Actualizar interfaz
            self.limpiar_campos()
            self.actualizar_lista_facturas()
            self.actualizar_resumen()
            self.statusBar().showMessage("Factura guardada correctamente", 3000)
    
    def validar_campos(self):
        """Validar los campos del formulario"""
        if not self.txt_descripcion.text().strip():
            QMessageBox.warning(self, "Validación", "La descripción es obligatoria")
            self.txt_descripcion.setFocus()
            return False
            
        if not self.txt_valor.text().strip() or float(self.txt_valor.text().replace(',', '.')) <= 0:
            QMessageBox.warning(self, "Validación", "El valor debe ser mayor a cero")
            self.txt_valor.setFocus()
            return False
            
        return True
    
    def limpiar_campos(self):
        """Limpiar los campos del formulario"""
        self.txt_descripcion.clear()
        self.txt_valor.clear()
        self.date_fecha.setDate(QDate.currentDate())
        self.cmb_tipo_gasto.setCurrentIndex(0)
    
    def actualizar_lista_facturas(self):
        """Actualizar la tabla de facturas"""
        # Desconectar temporalmente la señal para evitar múltiples llamadas
        try:
            self.tabla_facturas.itemChanged.disconnect(self.guardar_cambios_celda)
        except:
            pass
            
        # Cerrar cualquier editor activo
        self.tabla_facturas.closePersistentEditor(self.tabla_facturas.currentItem())
            
        self.tabla_facturas.setRowCount(len(self.facturas))
        
        # Establecer altura de fila para mejor legibilidad
        for i in range(len(self.facturas)):
            self.tabla_facturas.setRowHeight(i, 30)  # Aumentar altura de fila a 30 píxeles
            
        for i, factura in enumerate(self.facturas):
            # Columna 0 oculta para el ID de la factura
            item_id = QTableWidgetItem(str(factura.get('id', i)))
            item_id.setFlags(item_id.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabla_facturas.setItem(i, 0, item_id)
            
            # Fecha (editable)
            item_fecha = QTableWidgetItem(factura['fecha'])
            item_fecha.setFlags(item_fecha.flags() | Qt.ItemFlag.ItemIsEditable)
            self.tabla_facturas.setItem(i, 1, item_fecha)
            
            # Tipo (editable)
            item_tipo = QTableWidgetItem(factura['tipo'])
            item_tipo.setFlags(item_tipo.flags() | Qt.ItemFlag.ItemIsEditable)
            self.tabla_facturas.setItem(i, 2, item_tipo)
            
            # Descripción (editable)
            item_desc = QTableWidgetItem(factura['descripcion'])
            item_desc.setFlags(item_desc.flags() | Qt.ItemFlag.ItemIsEditable)
            self.tabla_facturas.setItem(i, 3, item_desc)
            
            # Valor (editable)
            valor_str = f"${float(factura['valor']):,.0f} COP".replace(',', '.')
            item_valor = QTableWidgetItem(valor_str)
            item_valor.setFlags(item_valor.flags() | Qt.ItemFlag.ItemIsEditable)
            self.tabla_facturas.setItem(i, 4, item_valor)
        
        # Reconectar la señal después de actualizar la tabla
        self.tabla_facturas.itemChanged.connect(self.guardar_cambios_celda)
    
    def actualizar_resumen(self):
        """Actualizar todos los resúmenes"""
        self.actualizar_resumen_diario()
        self.actualizar_resumen_mensual()
        self.actualizar_resumen_anual()
        self.actualizar_filtros()
    
    def actualizar_resumen_diario(self):
        """Actualizar el resumen diario"""
        # Verificar si los widgets de la interfaz están inicializados
        if not hasattr(self, 'date_resumen_diario') or not hasattr(self, 'texto_resumen_diario'):
            return
            
        try:
            fecha_seleccionada = self.date_resumen_diario.date()
            fecha_str = fecha_seleccionada.toString("dd/MM/yyyy")
            
            resumen = defaultdict(float)
            total = 0.0
            
            for factura in self.facturas:
                if factura['fecha'] == fecha_str:
                    tipo = factura['tipo']
                    valor = factura['valor']
                    resumen[tipo] += valor
                    total += valor
            
            # Formatear el resumen
            texto = f"Resumen de gastos para {fecha_str}\n\n"
            for tipo, monto in sorted(resumen.items()):
                texto += f"{tipo}: ${monto:,.0f} COP\n"
            
            texto += f"\nTotal del día: ${total:,.0f} COP"
            
            # Mostrar en el área de texto
            self.texto_resumen_diario.setPlainText(texto)
        except Exception as e:
            logger.error(f"Error al actualizar resumen diario: {str(e)}")
    
    def actualizar_resumen_mensual(self):
        """Actualizar el resumen mensual"""
        # Verificar si los widgets de la interfaz están inicializados
        if not all(hasattr(self, attr) for attr in ['combo_mes_resumen', 'combo_anio_resumen', 'texto_resumen_mensual']):
            return
            
        try:
            mes = self.combo_mes_resumen.currentIndex() + 1
            anio = int(self.combo_anio_resumen.currentText())
            
            resumen = defaultdict(float)
            total = 0.0
            
            for factura in self.facturas:
                try:
                    fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    if fecha.month == mes and fecha.year == anio:
                        tipo = factura['tipo']
                        valor = factura['valor']
                        resumen[tipo] += valor
                        total += valor
                except Exception as e:
                    logger.warning(f"Error al procesar factura: {str(e)}")
                    continue
            
            
            # Formatear el resumen
            nombre_mes = self.combo_mes_resumen.currentText()
            texto = f"Resumen de gastos para {nombre_mes} de {anio}\n\n"
            
            if resumen:
                for tipo, monto in sorted(resumen.items()):
                    porcentaje = (monto / total) * 100 if total > 0 else 0
                    texto += f"{tipo}: ${monto:,.0f} COP ({porcentaje:.1f}%)\n"
                
                texto += f"\nTotal del mes: ${total:,.0f} COP"
            else:
                texto += "No hay datos para mostrar en este período."
            
            # Mostrar en el área de texto
            self.texto_resumen_mensual.setPlainText(texto)
        except Exception as e:
            logger.error(f"Error al actualizar resumen mensual: {str(e)}")
    
    def actualizar_resumen_anual(self):
        """Actualizar el resumen anual"""
        # Verificar si los widgets de la interfaz están inicializados
        if not all(hasattr(self, attr) for attr in ['combo_anio_anual', 'texto_resumen_anual']):
            return
            
        try:
            anio = int(self.combo_anio_anual.currentText())
            
            resumen_mensual = defaultdict(lambda: defaultdict(float))
            resumen_anual = defaultdict(float)
            total_anual = 0.0
            
            for factura in self.facturas:
                try:
                    fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    if fecha.year == anio:
                        mes = fecha.month
                        tipo = factura['tipo']
                        valor = factura['valor']
                        
                        resumen_mensual[mes][tipo] += valor
                        resumen_anual[tipo] += valor
                        total_anual += valor
                except Exception as e:
                    logger.warning(f"Error al procesar factura: {str(e)}")
                    continue
            
            # Formatear el resumen
            texto = f"Resumen de gastos para el año {anio}\n\n"
            
            if resumen_anual:
                # Resumen por tipo de gasto
                texto += "=== Resumen por categoría ===\n"
                for tipo, monto in sorted(resumen_anual.items()):
                    porcentaje = (monto / total_anual) * 100 if total_anual > 0 else 0
                    texto += f"{tipo}: ${monto:,.0f} COP ({porcentaje:.1f}%)\n"
                
                # Resumen mensual
                texto += "\n=== Resumen mensual ===\n"
                meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
                
                for mes in range(1, 13):
                    total_mes = sum(resumen_mensual[mes].values())
                    if total_mes > 0:
                        porcentaje = (total_mes / total_anual) * 100 if total_anual > 0 else 0
                        texto += f"{meses[mes-1]}: ${total_mes:,.0f} COP ({porcentaje:.1f}%)\n"
                
                texto += f"\nTotal anual: ${total_anual:,.0f} COP"
            else:
                texto += "No hay datos para mostrar en este año."
            
            # Mostrar en el área de texto
            self.texto_resumen_anual.setPlainText(texto)
        except Exception as e:
            logger.error(f"Error al actualizar resumen anual: {str(e)}")
    
    def inicializar_filtros(self):
        """Inicializar los valores de los filtros"""
        # Obtener años únicos
        anios = set()
        for factura in self.facturas:
            try:
                fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                anios.add(fecha.year)
            except:
                continue
        
        # Ordenar años de mayor a menor
        anios_ordenados = sorted(anios, reverse=True)
        
        # Actualizar combo de años
        self.combo_filtro_anio.clear()
        self.combo_filtro_anio.addItem("Todos los años", None)
        for anio in anios_ordenados:
            self.combo_filtro_anio.addItem(str(anio), anio)
    
    def actualizar_filtros(self):
        """Actualizar los controles de filtro con los datos actuales"""
        self.inicializar_filtros()
        self.aplicar_filtros()
    
    def aplicar_filtros(self):
        """Aplicar los filtros seleccionados"""
        try:
            # Obtener fechas del rango
            qdate_desde = self.date_edit_desde.date()
            qdate_hasta = self.date_edit_hasta.date()
            
            # Convertir QDate a datetime para comparación
            fecha_desde = datetime(qdate_desde.year(), qdate_desde.month(), qdate_desde.day())
            fecha_hasta = datetime(qdate_hasta.year(), qdate_hasta.month(), qdate_hasta.day(), 23, 59, 59)
            
            anio = self.combo_filtro_anio.currentData()
            mes = self.combo_filtro_mes.currentIndex()  # 0 = Todos, 1-12 = meses
            dia = self.combo_filtro_dia.currentIndex()  # 0 = Todos, 1-31 = días
            tipo = self.combo_filtro_tipo.currentText() if self.combo_filtro_tipo.currentIndex() > 0 else None
            
            # Filtrar facturas
            facturas_filtradas = []
            for factura in self.facturas:
                try:
                    # Verificar si la factura tiene el formato de fecha esperado
                    if 'fecha' not in factura or not isinstance(factura['fecha'], str):
                        continue
                        
                    # Manejar diferentes formatos de fecha
                    try:
                        fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    except ValueError:
                        # Intentar con otro formato de fecha si es necesario
                        try:
                            fecha = datetime.strptime(factura['fecha'], '%Y-%m-%d')
                        except ValueError:
                            logger.warning(f"Formato de fecha no reconocido: {factura['fecha']}")
                            continue
                    
                    # Verificar rango de fechas
                    if fecha < fecha_desde or fecha > fecha_hasta:
                        continue
                    
                    # Verificar año
                    if anio is not None and fecha.year != anio:
                        continue
                    
                    # Verificar mes
                    if mes > 0 and fecha.month != mes:  # Si no es "Todos los meses"
                        continue
                    
                    # Verificar día
                    if dia > 0 and fecha.day != dia:  # Si no es "Todos los días"
                        continue
                    
                    # Verificar tipo
                    if tipo is not None and factura.get('tipo') != tipo:
                        continue
                    
                    facturas_filtradas.append(factura)
                except Exception as e:
                    logger.error(f"Error al procesar factura: {str(e)}", exc_info=True)
                    continue
            
            # Mostrar resultados solo si hay cambios para evitar actualizaciones innecesarias
            self.mostrar_resultados_filtrados(facturas_filtradas)
            
        except Exception as e:
            logger.error(f"Error en aplicar_filtros: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Se produjo un error al aplicar los filtros: {str(e)}")
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros"""
        # Restablecer fechas al rango por defecto (primero del mes actual hasta hoy)
        today = QDate.currentDate()
        first_day_of_month = QDate(today.year(), today.month(), 1)
        self.date_edit_desde.setDate(first_day_of_month)
        self.date_edit_hasta.setDate(today)
        
        # Restablecer otros filtros
        self.combo_filtro_anio.setCurrentIndex(0)
        self.combo_filtro_mes.setCurrentIndex(0)
        self.combo_filtro_dia.setCurrentIndex(0)
        self.combo_filtro_tipo.setCurrentIndex(0)
        self.aplicar_filtros()
    
    def mostrar_resultados_filtrados(self, facturas):
        """Mostrar las facturas filtradas en la tabla"""
        try:
            # Cerrar cualquier editor activo antes de actualizar
            if self.tabla_filtrada.currentItem() is not None:
                self.tabla_filtrada.closePersistentEditor(self.tabla_filtrada.currentItem())
                
            # Desconectar la señal temporalmente para evitar múltiples llamadas
            try:
                self.tabla_filtrada.itemChanged.disconnect(self.guardar_cambios_celda)
            except:
                pass  # La señal no estaba conectada
                
            # Configurar el número de filas
            self.tabla_filtrada.setRowCount(0)  # Limpiar la tabla
            self.tabla_filtrada.setRowCount(len(facturas) + 1)  # +1 para la fila de total
            
            # Establecer altura de fila para mejor legibilidad
            for i in range(len(facturas) + 1):
                self.tabla_filtrada.setRowHeight(i, 30)  # Aumentar altura de fila a 30 píxeles
            
            # Llenar la tabla con los datos de las facturas
            for i, factura in enumerate(facturas):
                if not isinstance(factura, dict):
                    continue
                
                try:
                    # Fecha (no editable)
                    fecha = factura.get('fecha', '')
                    fecha_item = QTableWidgetItem(str(fecha))
                    factura_id = factura.get('id')
                    if factura_id is not None:
                        fecha_item.setData(Qt.ItemDataRole.UserRole, factura_id)  # Store the ID for updates
                    self.tabla_filtrada.setItem(i, 0, fecha_item)
                    
                    # Tipo (editable con dropdown)
                    tipo = factura.get('tipo', '')
                    tipo_item = QTableWidgetItem(str(tipo))
                    tipo_item.setFlags(tipo_item.flags() | Qt.ItemFlag.ItemIsEditable | Qt.ItemFlag.ItemIsEnabled)
                    self.tabla_filtrada.setItem(i, 1, tipo_item)
                    
                    # Descripción (editable)
                    descripcion = factura.get('descripcion', '')
                    desc_item = QTableWidgetItem(str(descripcion))
                    desc_item.setFlags(desc_item.flags() | Qt.ItemFlag.ItemIsEditable)
                    self.tabla_filtrada.setItem(i, 2, desc_item)
                    
                    # Valor (editable)
                    valor = float(factura.get('valor', 0))
                    valor_formateado = f"${valor:,.0f} COP".replace(',', '.')
                    valor_item = QTableWidgetItem(valor_formateado)
                    valor_item.setData(Qt.ItemDataRole.UserRole + 1, valor)  # Store raw value for editing
                    valor_item.setFlags(valor_item.flags() | Qt.ItemFlag.ItemIsEditable)
                    self.tabla_filtrada.setItem(i, 3, valor_item)
                    
                except Exception as e:
                    logger.error(f"Error al mostrar factura {i}: {str(e)}", exc_info=True)
                    continue
            
            # Calcular total
            try:
                total = sum(float(factura.get('valor', 0)) for factura in facturas if isinstance(factura, dict))
            except Exception as e:
                logger.error(f"Error al calcular el total: {str(e)}")
                total = 0
            
            # Agregar fila de total
            total_row = len(facturas)
            
            # Celda vacía
            empty_item = QTableWidgetItem("")
            empty_item.setFlags(empty_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabla_filtrada.setItem(total_row, 0, empty_item)
            
            # Celda vacía
            empty_item2 = QTableWidgetItem("")
            empty_item2.setFlags(empty_item2.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.tabla_filtrada.setItem(total_row, 1, empty_item2)
            
            # Celda de TOTAL
            total_label = QTableWidgetItem("TOTAL:")
            total_label.setFlags(total_label.flags() & ~Qt.ItemFlag.ItemIsEditable)
            font = total_label.font()
            font.setBold(True)
            total_label.setFont(font)
            self.tabla_filtrada.setItem(total_row, 2, total_label)
            
            # Celda de valor total
            total_value = QTableWidgetItem(f"${total:,.0f} COP".replace(',', '.'))
            total_value.setFlags(total_value.flags() & ~Qt.ItemFlag.ItemIsEditable)
            total_value.setFont(font)
            self.tabla_filtrada.setItem(total_row, 3, total_value)
            
            # Resaltar la fila de total
            for col in range(self.tabla_filtrada.columnCount()):
                item = self.tabla_filtrada.item(total_row, col)
                if item:
                    if self.tema_oscuro:
                        item.setBackground(QColor(100, 100, 100))  # Gris oscuro para el tema oscuro
                        item.setForeground(QColor(255, 255, 255))  # Texto en blanco para mejor contraste
                    else:
                        item.setBackground(QColor(230, 230, 230))  # Gris claro para el tema claro
                        item.setForeground(QColor(0, 0, 0))  # Texto en negro
            
        except Exception as e:
            logger.error(f"Error en mostrar_resultados_filtrados: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Se produjo un error al mostrar los resultados: {str(e)}")
        finally:
            # Reconectar la señal al final, solo si no está ya conectada
            try:
                self.tabla_filtrada.itemChanged.disconnect()
            except:
                pass
            self.tabla_filtrada.itemChanged.connect(self.guardar_cambios_celda)

    def _mostrar_vista_previa(self, facturas, tipo_archivo):
        """Mostrar una vista previa de las facturas a importar"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Vista Previa - {tipo_archivo}")
        dialog.resize(800, 500)
        
        layout = QVBoxLayout(dialog)
        
        # Crear tabla para la vista previa
        tabla = QTableWidget(dialog)
        tabla.setColumnCount(5)
        tabla.setHorizontalHeaderLabels(["Fecha", "Tipo", "Descripción", "Valor (COP)", "Estado"])
        tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        # Mostrar hasta 10 filas como vista previa
        num_filas = min(10, len(facturas))
        tabla.setRowCount(num_filas)
        
        # Llenar la tabla con datos de vista previa
        for i in range(num_filas):
            factura = facturas[i]
            tabla.setItem(i, 0, QTableWidgetItem(factura.get('fecha', '')))
            tabla.setItem(i, 1, QTableWidgetItem(factura.get('tipo', '')))
            tabla.setItem(i, 2, QTableWidgetItem(factura.get('descripcion', '')))
            tabla.setItem(i, 3, QTableWidgetItem(f"${factura.get('valor', 0):,.0f} COP".replace(',', '.')))
            tabla.setItem(i, 4, QTableWidgetItem("✅ Válido"))
        
        # Mostrar resumen
        total_facturas = len(facturas)
        resumen = QLabel(f"Mostrando {num_filas} de {total_facturas} facturas encontradas.")
        
        # Botones
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        
        layout.addWidget(QLabel("Vista previa de las primeras facturas a importar:"))
        layout.addWidget(tabla)
        layout.addWidget(resumen)
        layout.addWidget(btn_box)
        
        return dialog.exec() == QDialog.DialogCode.Accepted
    
    def _procesar_importacion(self, facturas_importadas, tipo_archivo):
        """Procesar la importación con diálogo de progreso"""
        # Crear diálogo de progreso
        progress = QProgressDialog(
            f"Importando {len(facturas_importadas)} facturas desde {tipo_archivo}...",
            "Cancelar", 0, len(facturas_importadas), self)
        progress.setWindowTitle("Importando facturas")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(1000)  # Mostrar después de 1 segundo
        
        # Mostrar diálogo de confirmación con vista previa
        if not self._mostrar_vista_previa(facturas_importadas, tipo_archivo):
            return False
        
        # Preguntar al usuario si desea sobrescribir o agregar
        reply = QMessageBox.question(
            self,
            f'Importar {tipo_archivo}',
            f'¿Desea sobrescribir las facturas existentes con las {len(facturas_importadas)} facturas del archivo?\n' \
            '"No" agregará las facturas a las existentes.',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel
        )
        
        if reply == QMessageBox.StandardButton.Cancel:
            return False
        
        # Guardar copia de seguridad de los datos actuales
        facturas_originales = self.facturas.copy()
        
        try:
            # Actualizar la barra de estado
            self.statusBar().showMessage(f"Importando {len(facturas_importadas)} facturas...")
            
            # Procesar la importación con barra de progreso
            for i, factura in enumerate(facturas_importadas, 1):
                if progress.wasCanceled():
                    break
                    
                progress.setValue(i)
                QApplication.processEvents()  # Mantener la interfaz responsiva
                
            # Aplicar los cambios según la opción seleccionada
            if reply == QMessageBox.StandardButton.Yes:
                self.facturas = facturas_importadas
            else:
                self.facturas.extend(facturas_importadas)
            
            # Guardar los datos
            if not self.guardar_datos():
                raise Exception("No se pudieron guardar los datos")
            
            # Actualizar la interfaz
            self.actualizar_tabla_facturas()
            self.actualizar_resumenes()
            
            # Mostrar notificación de éxito
            self.statusBar().showMessage(
                f"Se importaron {len(facturas_importadas)} facturas correctamente.", 
                5000  # Mostrar por 5 segundos
            )
            
            # Mostrar mensaje de éxito con más detalles
            QMessageBox.information(
                self, 
                "Importación exitosa", 
                f"Se importaron {len(facturas_importadas)} facturas desde el archivo {tipo_archivo}.\n\n"
                f"Total de facturas actuales: {len(self.facturas)}"
            )
            
            return True
            
        except Exception as e:
            # En caso de error, restaurar los datos originales
            self.facturas = facturas_originales
            error_msg = (
                f"Error al importar las facturas:\n\n"
                f"Error: {str(e)}\n\n"
                f"Se han restaurado los datos originales."
            )
            logger.error(f"Error en importación: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error en la importación", error_msg)
            return False
    
    def importar_desde_csv(self):
        """Importar facturas desde un archivo CSV"""
        # Abrir diálogo para seleccionar archivo
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo CSV",
            "",
            "Archivos CSV (*.csv);;Todos los archivos (*)"
        )
        
        if not file_path:
            return  # Usuario canceló el diálogo
        
        # Mostrar indicador de carga
        loading_box = QMessageBox(self)
        loading_box.setWindowTitle("Cargando archivo")
        loading_box.setText("Procesando archivo CSV, por favor espere...")
        loading_box.setStandardButtons(QMessageBox.StandardButton.NoButton)
        loading_box.show()
        QApplication.processEvents()
        
        try:
            facturas_importadas = []
            errores = []
            total_lineas = 0
            
            # Primera pasada: contar líneas para la barra de progreso
            with open(file_path, 'r', encoding='utf-8') as f:
                total_lineas = sum(1 for _ in f) - 1  # Restar 1 por el encabezado
            
            # Crear diálogo de progreso para la lectura
            progress = QProgressDialog("Leyendo archivo CSV...", "Cancelar", 0, total_lineas, self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            
            # Segunda pasada: procesar el archivo
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                lineas_procesadas = 0
                
                # Verificar que el CSV tenga las columnas necesarias
                required_columns = {'fecha', 'tipo', 'descripcion', 'valor'}
                if not required_columns.issubset(reader.fieldnames):
                    progress.close()
                    QMessageBox.critical(
                        self, 
                        "Error de formato", 
                        f"El archivo CSV debe contener las columnas: {', '.join(required_columns)}\n\n"
                        f"Columnas encontradas: {', '.join(reader.fieldnames) if reader.fieldnames else 'Ninguna'}"
                    )
                    return
                
                for row in reader:
                    if progress.wasCanceled():
                        return
                        
                    lineas_procesadas += 1
                    progress.setValue(lineas_procesadas)
                    QApplication.processEvents()
                    
                    try:
                        # Validar y convertir los datos
                        fecha = row['fecha'].strip()
                        tipo = row['tipo'].strip()
                        descripcion = row['descripcion'].strip()
                        
                        # Validar campos obligatorios
                        if not all([fecha, tipo, descripcion, row['valor']]):
                            errores.append(f"Fila {lineas_procesadas + 1}: Faltan campos obligatorios")
                            continue
                        
                        try:
                            # Manejar diferentes formatos de valor (con comas, puntos, etc.)
                            valor_str = str(row['valor']).replace('$', '').replace(',', '').strip()
                            valor = float(valor_str)
                            if valor <= 0:
                                raise ValueError("El valor debe ser mayor a cero")
                        except (ValueError, TypeError) as ve:
                            errores.append(f"Fila {lineas_procesadas + 1}: Valor inválido - {str(ve)}")
                            continue
                        
                        # Validar formato de fecha (DD/MM/YYYY)
                        try:
                            fecha_dt = datetime.strptime(fecha, '%d/%m/%Y')
                            fecha = fecha_dt.strftime('%d/%m/%Y')  # Estandarizar formato
                        except ValueError:
                            errores.append(f"Fila {lineas_procesadas + 1}: Formato de fecha inválido (debe ser DD/MM/YYYY)")
                            continue
                        
                        # Crear objeto de factura
                        factura = {
                            'fecha': fecha,
                            'tipo': tipo,
                            'descripcion': descripcion,
                            'valor': valor,
                            'timestamp': datetime.now().isoformat()
                        }
                        facturas_importadas.append(factura)
                        
                    except Exception as e:
                        error_msg = f"Fila {lineas_procesadas + 1}: {str(e)}"
                        logger.error(f"Error al procesar fila del CSV: {row}. Error: {error_msg}")
                        errores.append(error_msg)
                        continue
            
            progress.close()
            loading_box.close()
            
            # Mostrar advertencias si hay errores
            if errores:
                error_dialog = QDialog(self)
                error_dialog.setWindowTitle("Advertencias de importación")
                error_dialog.resize(600, 400)
                
                layout = QVBoxLayout()
                
                # Mostrar resumen
                resumen = QLabel(
                    f"Se encontraron {len(errores)} advertencias durante la importación.\n"
                    f"Se importaron {len(facturas_importadas)} facturas correctamente."
                )
                
                # Área de texto para mostrar los errores
                error_text = QTextEdit()
                error_text.setReadOnly(True)
                error_text.setPlainText("\n".join(errores))
                
                # Botón para continuar
                btn_continuar = QPushButton("Continuar con la importación")
                btn_continuar.clicked.connect(error_dialog.accept)
                
                # Botón para cancelar
                btn_cancelar = QPushButton("Cancelar importación")
                btn_cancelar.clicked.connect(error_dialog.reject)
                
                # Layout de botones
                btn_layout = QHBoxLayout()
                btn_layout.addStretch()
                btn_layout.addWidget(btn_continuar)
                btn_layout.addWidget(btn_cancelar)
                
                # Agregar widgets al layout
                layout.addWidget(resumen)
                layout.addWidget(QLabel("Detalles de las advertencias:"))
                layout.addWidget(error_text)
                layout.addLayout(btn_layout)
                
                error_dialog.setLayout(layout)
                
                if error_dialog.exec() != QDialog.DialogCode.Accepted:
                    return False
            
            if not facturas_importadas:
                QMessageBox.warning(
                    self, 
                    "Ninguna factura válida", 
                    "No se encontraron facturas válidas en el archivo."
                )
                return
            
            # Procesar la importación con la nueva función de ayuda
            return self._procesar_importacion(facturas_importadas, "CSV")
            
        except Exception as e:
            loading_box.close()
            error_msg = (
                f"Error al procesar el archivo CSV:\n\n"
                f"Error: {str(e)}\n\n"
                f"Asegúrese de que el archivo no esté abierto en otro programa y que tenga el formato correcto."
            )
            logger.error(f"Error al importar desde CSV: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error en la importación", error_msg)
            return False
    
    def importar_desde_excel(self):
        """Importar facturas desde un archivo Excel"""
        # Abrir diálogo para seleccionar archivo
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel",
            "",
            "Archivos Excel (*.xlsx *.xls);;Todos los archivos (*)"
        )
        
        if not file_path:
            return  # Usuario canceló el diálogo
        
        # Mostrar indicador de carga
        loading_box = QMessageBox(self)
        loading_box.setWindowTitle("Cargando archivo")
        loading_box.setText("Procesando archivo Excel, por favor espere...")
        loading_box.setStandardButtons(QMessageBox.StandardButton.NoButton)
        loading_box.show()
        QApplication.processEvents()
        
        try:
            # Cargar el libro de trabajo de Excel
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Crear diálogo para seleccionar hoja
            sheet_dialog = QDialog(self)
            sheet_dialog.setWindowTitle("Seleccionar hoja")
            sheet_dialog.setMinimumWidth(300)
            
            layout = QVBoxLayout()
            
            # Mostrar resumen
            resumen = QLabel("Seleccione la hoja que contiene los datos:")
            
            # Lista de hojas disponibles
            sheet_list = QListWidget()
            sheet_list.addItems(wb.sheetnames)
            sheet_list.setCurrentRow(0)  # Seleccionar la primera hoja por defecto
            layout.addWidget(sheet_list)
            
            # Botones
            btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            btn_box.accepted.connect(sheet_dialog.accept)
            btn_box.rejected.connect(sheet_dialog.reject)
            layout.addWidget(btn_box)
            
            sheet_dialog.setLayout(layout)
            
            if sheet_dialog.exec() != QDialog.DialogCode.Accepted:
                loading_box.close()
                return
            
            # Obtener la hoja seleccionada
            selected_sheet = sheet_list.currentItem().text()
            sheet = wb[selected_sheet]
            
            # Actualizar mensaje de carga
            loading_box.setText(f"Procesando hoja: {selected_sheet}...")
            QApplication.processEvents()
            
            # Obtener los encabezados
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
            
            # Verificar que los encabezados requeridos estén presentes
            required_columns = ['fecha', 'tipo', 'descripcion', 'valor']
            
            # Mapear los índices de las columnas
            column_indices = {}
            missing_columns = []
            
            for col in required_columns:
                try:
                    column_indices[col] = headers.index(col)
                except ValueError:
                    missing_columns.append(col)
            
            if missing_columns:
                loading_box.close()
                QMessageBox.critical(
                    self, 
                    "Error de formato", 
                    f"El archivo Excel debe contener las columnas: {', '.join(required_columns)}.\n"
                    f"Columnas faltantes: {', '.join(missing_columns)}\n\n"
                    f"Columnas encontradas: {', '.join([h for h in headers if h]) or 'Ninguna'}"
                )
                return
            
            # Contar filas para la barra de progreso
            total_rows = sheet.max_row - 1  # Restar la fila de encabezados
            
            # Crear diálogo de progreso para la lectura
            progress = QProgressDialog("Leyendo archivo Excel...", "Cancelar", 0, total_rows, self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            
            # Leer las facturas
            facturas_importadas = []
            errores = []
            
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):  # Empezar desde la fila 2
                if progress.wasCanceled():
                    loading_box.close()
                    return
                
                progress.setValue(i-1)  # Ajustar para empezar desde 0
                QApplication.processEvents()
                
                try:
                    # Obtener los valores de las celdas
                    fecha = str(row[column_indices['fecha']]).strip() if row[column_indices['fecha']] else ""
                    tipo = str(row[column_indices['tipo']]).strip() if row[column_indices['tipo']] else ""
                    descripcion = str(row[column_indices['descripcion']]).strip() if row[column_indices['descripcion']] else ""
                    valor_celda = row[column_indices['valor']]
                    
                    # Validar campos obligatorios
                    if not all([fecha, tipo, descripcion, valor_celda is not None]):
                        errores.append(f"Fila {i}: Faltan campos obligatorios")
                        continue
                    
                    try:
                        # Manejar diferentes formatos de valor
                        valor_str = str(valor_celda).replace('$', '').replace(',', '').strip()
                        valor = float(valor_str)
                        if valor <= 0:
                            raise ValueError("El valor debe ser mayor a cero")
                    except (ValueError, TypeError) as ve:
                        errores.append(f"Fila {i}: Valor inválido - {str(ve)}")
                        continue
                    
                    # Validar formato de fecha (DD/MM/YYYY)
                    try:
                        fecha_dt = datetime.strptime(fecha, '%d/%m/%Y')
                        fecha = fecha_dt.strftime('%d/%m/%Y')  # Estandarizar formato
                    except ValueError:
                        errores.append(f"Fila {i}: Formato de fecha inválido (debe ser DD/MM/YYYY)")
                        continue
                    
                    # Crear objeto de factura
                    factura = {
                        'fecha': fecha,
                        'tipo': tipo,
                        'descripcion': descripcion,
                        'valor': valor,
                        'timestamp': datetime.now().isoformat()
                    }
                    facturas_importadas.append(factura)
                    
                except Exception as e:
                    error_msg = f"Fila {i}: {str(e)}"
                    logger.error(f"Error al procesar fila del Excel: {row}. Error: {error_msg}")
                    errores.append(error_msg)
                    continue
            
            progress.close()
            loading_box.close()
            
            # Mostrar advertencias si hay errores
            if errores:
                error_dialog = QDialog(self)
                error_dialog.setWindowTitle("Advertencias de importación")
                error_dialog.resize(600, 400)
                
                layout = QVBoxLayout()
                
                # Mostrar resumen
                resumen = QLabel(
                    f"Se encontraron {len(errores)} advertencias durante la importación.\n"
                    f"Se importaron {len(facturas_importadas)} facturas correctamente."
                )
                
                # Área de texto para mostrar los errores
                error_text = QTextEdit()
                error_text.setReadOnly(True)
                error_text.setPlainText("\n".join(errores))
                
                # Botón para continuar
                btn_continuar = QPushButton("Continuar con la importación")
                btn_continuar.clicked.connect(error_dialog.accept)
                
                # Botón para cancelar
                btn_cancelar = QPushButton("Cancelar importación")
                btn_cancelar.clicked.connect(error_dialog.reject)
                
                # Layout de botones
                btn_layout = QHBoxLayout()
                btn_layout.addStretch()
                btn_layout.addWidget(btn_continuar)
                btn_layout.addWidget(btn_cancelar)
                
                # Agregar widgets al layout
                layout.addWidget(resumen)
                layout.addWidget(QLabel("Detalles de las advertencias:"))
                layout.addWidget(error_text)
                layout.addLayout(btn_layout)
                
                error_dialog.setLayout(layout)
                
                if error_dialog.exec() != QDialog.DialogCode.Accepted:
                    return False
            
            if not facturas_importadas:
                QMessageBox.warning(
                    self, 
                    "Ninguna factura válida", 
                    "No se encontraron facturas válidas en el archivo."
                )
                return
            
            # Procesar la importación con la nueva función de ayuda
            return self._procesar_importacion(facturas_importadas, "Excel")
            
        except Exception as e:
            loading_box.close()
            error_msg = (
                f"Error al procesar el archivo Excel:\n\n"
                f"Error: {str(e)}\n\n"
                f"Asegúrese de que el archivo no esté abierto en otro programa y que tenga el formato correcto."
            )
            logger.error(f"Error al importar desde Excel: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error en la importación", error_msg)
            return False
    
    def importar_desde_json(self):
        """Importar facturas desde un archivo JSON con diálogo de progreso y validación"""  
        # Abrir diálogo para seleccionar archivo
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo JSON",
            "",
            "Archivos JSON (*.json);;Todos los archivos (*)"
        )
    
        if not file_path:
            return  # Usuario canceló el diálogo
        
        # Mostrar indicador de carga
        loading_box = QMessageBox(self)
        loading_box.setWindowTitle("Cargando archivo")
        loading_box.setText("Procesando archivo JSON, por favor espere...")
        loading_box.setStandardButtons(QMessageBox.StandardButton.NoButton)
        loading_box.show()
        QApplication.processEvents()
        
        try:
            # Leer el archivo JSON
            with open(file_path, 'r', encoding='utf-8') as f:
                facturas_importadas = json.load(f)
            
            loading_box.close()
            
            if not isinstance(facturas_importadas, list):
                QMessageBox.critical(
                    self,
                    "Error",
                    "El archivo JSON debe contener una lista de facturas."
                )
                return
                
            # Validar facturas
            facturas_validas = []
            errores = []
            
            # Crear diálogo de progreso
            progress = QProgressDialog(
                f"Validando {len(facturas_importadas)} facturas...",
                "Cancelar", 0, len(facturas_importadas), self)
            progress.setWindowTitle("Validando facturas")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(500)  # Mostrar después de 500ms
            
            # Validar cada factura
            for i, factura in enumerate(facturas_importadas):
                if progress.wasCanceled():
                    QMessageBox.information(self, "Importación cancelada", 
                                        "La importación ha sido cancelada por el usuario.")
                    return
                    
                progress.setValue(i)
                progress.setLabelText(f"Validando factura {i+1} de {len(facturas_importadas)}...")
                QApplication.processEvents()
                
                try:
                    # Validar campos obligatorios
                    if not all(key in factura for key in ['tipo', 'descripcion', 'valor', 'fecha']):
                        errores.append(f"Factura {i+1}: Faltan campos obligatorios")
                        continue
                    
                    # Validar y convertir tipos de datos
                    try:
                        factura['valor'] = float(factura['valor'])
                        if factura['valor'] <= 0:
                            raise ValueError("El valor debe ser mayor a cero")
                    except (ValueError, TypeError) as e:
                        errores.append(f"Factura {i+1}: Valor inválido - {str(e)}")
                        continue
                    
                    # Validar fecha
                    try:
                        datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    except (ValueError, TypeError):
                        errores.append(f"Factura {i+1}: Formato de fecha inválido. Use DD/MM/AAAA")
                        continue
                    
                    facturas_validas.append(factura)
                    
                except Exception as e:
                    errores.append(f"Factura {i+1}: Error inesperado - {str(e)}")
                    continue
                    
            progress.close()
            
            # Mostrar advertencias si hay errores
            if errores:
                error_dialog = QDialog(self)
                error_dialog.setWindowTitle("Advertencias de importación")
                error_dialog.resize(600, 400)
                
                layout = QVBoxLayout()
                
                label = QLabel(f"Se encontraron {len(errores)} advertencias al importar {len(facturas_importadas)} facturas. "
                            f"Se importarán {len(facturas_validas)} facturas correctas.")
                label.setWordWrap(True)
                layout.addWidget(label)
                
                text_area = QTextEdit()
                text_area.setReadOnly(True)
                text_area.setPlainText("\n".join(errores))
                layout.addWidget(text_area)
                
                button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
                button_box.accepted.connect(error_dialog.accept)
                layout.addWidget(button_box)
                
                error_dialog.setLayout(layout)
                error_dialog.exec()
                
                if not facturas_validas:
                    return  # No hay facturas válidas para importar
            
            # Usar el método _procesar_importacion para manejar la importación
            if facturas_validas:
                # Guardar copia de seguridad
                facturas_originales = self.facturas.copy()
                
                try:
                    # Agregar facturas importadas
                    self.facturas.extend(facturas_validas)
                    
                    # Actualizar interfaz
                    self.actualizar_lista_facturas()
                    
                    # Actualizar resumen solo si el widget existe
                    if hasattr(self, 'actualizar_resumen'):
                        try:
                            self.actualizar_resumen()
                        except AttributeError as e:
                            logger.warning(f"No se pudo actualizar el resumen: {str(e)}")
                    
                    # Guardar datos
                    self.guardar_datos()
                    
                    QMessageBox.information(
                        self,
                        "Importación exitosa",
                        f"Se importaron {len(facturas_validas)} facturas desde {file_path}."
                    )
                    logger.info(f"Se importaron {len(facturas_validas)} facturas desde {file_path}")
                    
                except Exception as e:
                    # Revertir cambios en caso de error
                    self.facturas = facturas_originales
                    self.actualizar_lista_facturas()
                    if hasattr(self, 'actualizar_resumen'):
                        try:
                            self.actualizar_resumen()
                        except:
                            pass
                    
                    error_msg = f"Error al procesar las facturas: {str(e)}"
                    logger.error(error_msg, exc_info=True)
                    QMessageBox.critical(self, "Error", error_msg)
        
        except json.JSONDecodeError:
            QMessageBox.critical(self, "Error", "El archivo seleccionado no es un JSON válido.")
        except Exception as e:
            error_msg = f"Error al importar desde JSON: {str(e)}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Error", error_msg)
    
    def cargar_datos(self, actualizar_ui=True):
        """Cargar datos desde la base de datos
        
        Args:
            actualizar_ui (bool): Si es True, actualiza la interfaz de usuario
        """
        try:
            # Obtener todos los tipos de gasto de la base de datos
            self.tipos_gasto = self.db.obtener_tipos_gasto()
            
            # Obtener todas las facturas de la base de datos
            self.facturas = self.db.obtener_facturas()
            
            # Actualizar la interfaz si está solicitado y los componentes existen
            if actualizar_ui:
                if hasattr(self, 'tabla_facturas'):
                    # Actualizar el delegado de la columna de tipo con los tipos de gasto actualizados
                    self.tabla_facturas.setItemDelegateForColumn(2, TipoGastoDelegate(self.tabla_facturas, self.tipos_gasto))
                    self.actualizar_lista_facturas()
                if hasattr(self, 'actualizar_resumen'):
                    self.actualizar_resumen()
            
            logger.info(f"Se cargaron {len(self.facturas)} facturas y {len(self.tipos_gasto)} tipos de gasto desde la base de datos")
            return True
            
        except Exception as e:
            error_msg = f"Error al cargar los datos de la base de datos: {str(e)}"
            logger.error(error_msg, exc_info=True)
            if hasattr(self, 'isVisible'):  # Solo mostrar mensaje si la ventana está visible
                QMessageBox.critical(self, "Error", error_msg)
            self.facturas = []
            self.tipos_gasto = []
            return False
            
    def guardar_datos(self):
        """Guardar datos en la base de datos"""
        try:
            # Guardar todas las facturas en la base de datos
            if hasattr(self, 'facturas'):
                # Obtener las facturas actuales de la base de datos
                facturas_actuales = {f['id']: f for f in self.db.obtener_facturas()}
                
                # Actualizar o insertar facturas
                for factura in self.facturas:
                    if 'id' in factura and factura['id'] in facturas_actuales:
                        # Actualizar factura existente
                        self.db.actualizar_factura(
                            factura_id=factura['id'],
                            fecha=factura['fecha'],
                            tipo=factura['tipo'],
                            descripcion=factura['descripcion'],
                            valor=factura['valor']
                        )
                        # Eliminar de facturas_actuales para marcar como procesada
                        facturas_actuales.pop(factura['id'], None)
                    else:
                        # Insertar nueva factura
                        factura_id = self.db.agregar_factura(
                            fecha=factura['fecha'],
                            tipo=factura['tipo'],
                            descripcion=factura['descripcion'],
                            valor=factura['valor']
                        )
                        # Actualizar el ID en la factura local
                        factura['id'] = factura_id
                
                # Eliminar facturas que ya no están en self.facturas
                for factura_id in facturas_actuales:
                    self.db.eliminar_factura(factura_id)
                
                logger.info(f"Se guardaron {len(self.facturas)} facturas en la base de datos")
                return True
            return False
            
        except Exception as e:
            error_msg = f"Error al guardar los datos en la base de datos: {str(e)}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Error", error_msg)
            return False
    
    def exportar_a_excel(self):
        """Exportar los datos a un archivo Excel con formato de tabla"""
        if not self.facturas:
            QMessageBox.warning(self, "Exportar a Excel", "No hay datos para exportar.")
            return
        
        # Obtener la configuración
        config = get_config()
        last_dir = config['APP'].get('last_export_dir', str(Path.home() / 'Documents'))
        
        # Nombre de archivo predeterminado con la fecha actual
        default_filename = f"Facturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        default_path = str(Path(last_dir) / default_filename)
        
        # Solicitar al usuario la ubicación para guardar el archivo
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar como",
            default_path,
            "Archivos Excel (*.xlsx);;Todos los archivos (*)"
        )
        
        if not file_path:
            return  # Usuario canceló el diálogo
        
        # Asegurarse de que el archivo tenga la extensión .xlsx
        if not file_path.lower().endswith('.xlsx'):
            file_path += '.xlsx'
            
        try:
            # Crear un nuevo libro de Excel
            wb = Workbook()
            
            # Eliminar la hoja por defecto si no tiene datos
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            money_format = '#,##0.00" COP"'
            
            # 1. Hoja de Detalle (todas las facturas)
            ws_detalle = wb.create_sheet("Todas las Facturas")
            
            # Encabezados
            headers = ["Fecha", "Tipo", "Descripción", "Valor (COP)"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_detalle.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos
            for row_num, factura in enumerate(self.facturas, 2):
                ws_detalle.cell(row=row_num, column=1, value=factura['fecha'])
                ws_detalle.cell(row=row_num, column=2, value=factura['tipo'])
                ws_detalle.cell(row=row_num, column=3, value=factura['descripcion'])
                
                # Formato de moneda para la columna de valor
                cell_valor = ws_detalle.cell(row=row_num, column=4, value=float(factura['valor']))
                cell_valor.number_format = money_format
            
            # Crear tabla
            table = Table(displayName="TablaDetalle", ref=f"A1:D{len(self.facturas) + 1}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws_detalle.add_table(table)
            
            # Ajustar ancho de columnas
            for column in ws_detalle.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        value_length = len(str(cell.value).encode('utf-8'))
                        if value_length > max_length:
                            max_length = value_length
                    except:
                        pass
                adjusted_width = (max_length + 4) * 1.1
                ws_detalle.column_dimensions[column_letter].width = min(adjusted_width, 40)
            
            # 2. Agrupar facturas por mes y año
            facturas_por_mes = {}
            for factura in self.facturas:
                try:
                    # Intentar parsear la fecha en formato dd/mm/yyyy
                    fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    mes_anio = fecha.strftime('%Y-%m')  # Formato: YYYY-MM

                    if mes_anio not in facturas_por_mes:
                        facturas_por_mes[mes_anio] = []
                    facturas_por_mes[mes_anio].append(factura)
                except ValueError as e:
                    logger.warning(f"No se pudo parsear la fecha: {factura['fecha']}. Error: {e}")
                    continue
            
            # 3. Crear una hoja para cada mes
            for mes_anio, facturas_mes in facturas_por_mes.items():
                try:
                    # Crear nombre de hoja en formato 'YYYY-MM Mes' (ej. '2025-08 Agosto')
                    fecha = datetime.strptime(mes_anio, '%Y-%m')
                    nombre_mes = f"{fecha.strftime('%Y-%m')} {fecha.strftime('%B').capitalize()}"

                    # Limitar el nombre de la hoja a 31 caracteres (límite de Excel)
                    nombre_hoja = nombre_mes[:31]

                    # Crear hoja para el mes
                    ws_mes = wb.create_sheet(nombre_hoja)

                    # Título del mes
                    titulo_mes = f"Facturas de {fecha.strftime('%B').capitalize()} {fecha.year}"
                    ws_mes.append([titulo_mes])
                    ws_mes.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                    titulo_cell = ws_mes.cell(row=1, column=1)
                    titulo_cell.font = Font(size=14, bold=True)
                    titulo_cell.alignment = Alignment(horizontal='center')

                    # Resumen del mes
                    total_mes = sum(f['valor'] for f in facturas_mes)
                    ws_mes.append([f"Total del mes: ${total_mes:,.0f} COP"])
                    ws_mes.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
                    total_cell = ws_mes.cell(row=2, column=1)
                    total_cell.font = Font(bold=True)

                    # Espacio antes de la tabla
                    ws_mes.append([])

                    # Encabezados de la tabla
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_mes.cell(row=4, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill

                    # Datos de las facturas del mes
                    # Escribir los datos de las facturas
                    for row_num, factura in enumerate(facturas_mes, 5):
                        ws_mes.cell(row=row_num, column=1, value=factura['fecha'])
                        ws_mes.cell(row=row_num, column=2, value=factura['tipo'])
                        ws_mes.cell(row=row_num, column=3, value=factura['descripcion'])

                        # Formato de moneda para la columna de valor
                        cell_valor = ws_mes.cell(row=row_num, column=4, value=float(factura['valor']))
                        cell_valor.number_format = money_format
                    
                    # Crear tabla para el mes después de escribir todos los datos
                    if facturas_mes:  # Solo crear tabla si hay datos
                        try:
                            # Calcular la última fila de datos (5 + len - 1 ya que empezamos en 5)
                            last_row = 4 + len(facturas_mes)
                            table_ref = f"A4:D{last_row}"  # Rango desde A4 hasta D{última fila}
                            
                            # Crear un nombre de tabla único y válido para Excel
                            table_name = f"Tabla_{mes_anio.replace('-', '_')}"
                            
                            # Asegurarse de que no haya tablas con el mismo nombre
                            if table_name in ws_mes._tables:
                                del ws_mes._tables[table_name]
                            
                            # Crear la tabla
                            table = Table(displayName=table_name, ref=table_ref)
                            
                            # Configurar el estilo de la tabla
                            style = TableStyleInfo(
                                name="TableStyleMedium9",
                                showFirstColumn=False,
                                showLastColumn=False,
                                showRowStripes=True,
                                showColumnStripes=False
                            )
                            table.tableStyleInfo = style
                            
                            # Agregar la tabla a la hoja
                            ws_mes.add_table(table)
                            
                        except Exception as e:
                            logger.warning(f"No se pudo crear la tabla para el mes {mes_anio}: {str(e)}")
                            # Continuar con la ejecución a pesar del error en la tabla
                        
                        # Ajustar ancho de columnas
                        # Primero, obtener el rango de celdas que no están fusionadas
                        data_rows = []
                        for row in ws_mes.iter_rows(min_row=4):  # Empezar después de los títulos
                            if all(cell.value is not None for cell in row):
                                data_rows.append(row)
                        
                        # Si no hay filas de datos, saltar el ajuste
                        if not data_rows:
                            continue
                            
                        # Obtener el número de columnas
                        num_columns = len(data_rows[0])
                        
                        # Ajustar el ancho de cada columna
                        for col_idx in range(num_columns):
                            try:
                                # Obtener la letra de la columna
                                column_letter = get_column_letter(col_idx + 1)
                                max_length = 0
                                
                                # Ajustar para el encabezado (fila 4)
                                header_cell = ws_mes.cell(row=4, column=col_idx + 1)
                                if header_cell and header_cell.value:
                                    max_length = len(str(header_cell.value))
                                
                                # Especificar un ancho mínimo para la columna de valor (columna D)
                                if column_letter == 'D':
                                    max_length = max(max_length, 15)  # Mínimo para valores monetarios
                                
                                # Revisar las celdas de datos
                                for row in data_rows:
                                    cell = row[col_idx]
                                    if cell and cell.value is not None:
                                        # Para celdas con formato de moneda
                                        if hasattr(cell, 'number_format') and 'COP' in str(cell.number_format):
                                            formatted_value = f"{float(cell.value):,.2f} COP"
                                            value_length = len(formatted_value)
                                        else:
                                            value_length = len(str(cell.value))
                                        
                                        if value_length > max_length:
                                            max_length = value_length
                                
                                # Ajustar el ancho con espacio adicional
                                if max_length > 0:
                                    adjusted_width = (max_length + 2) * 1.2
                                    # Limitar el ancho máximo a 50 caracteres
                                    ws_mes.column_dimensions[column_letter].width = min(adjusted_width, 50)
                                    
                                    # Asegurar que la columna de valor tenga un ancho mínimo
                                    if column_letter == 'D':
                                        ws_mes.column_dimensions[column_letter].width = max(
                                            ws_mes.column_dimensions[column_letter].width, 15
                                        )
                                        
                            except Exception as e:
                                logger.warning(f"No se pudo ajustar el ancho de la columna {col_idx + 1}: {str(e)}")
                                continue
                
                except Exception as e:
                    logger.error(f"Error al crear la hoja para el mes {mes_anio}: {str(e)}")
                    continue

            
            # 4. Hoja de Resumen por Tipo de Gasto
            ws_resumen = wb.create_sheet("Resumen por Tipo")
            
            # Calcular totales por tipo
            total_por_tipo = defaultdict(float)
            for factura in self.facturas:
                total_por_tipo[factura['tipo']] += factura['valor']
            
            # Ordenar por valor descendente
            total_por_tipo = dict(sorted(total_por_tipo.items(), key=lambda x: x[1], reverse=True))
            
            # Escribir encabezados
            ws_resumen.append(["Tipo de Gasto", "Total (COP)", "Porcentaje"])
            
            # Calcular total general para porcentajes
            total_general = sum(total_por_tipo.values())
            
            # Escribir datos
            for tipo, total in total_por_tipo.items():
                if total_general > 0:
                    porcentaje = (total / total_general) * 100
                    ws_resumen.append([tipo, total, porcentaje/100])
                else:
                    ws_resumen.append([tipo, total, 0])
            
            # Aplicar formato a la tabla de resumen
            table_resumen = Table(displayName="TablaResumenTipo", ref=f"A1:C{len(total_por_tipo) + 1}")
            style_resumen = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                         showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table_resumen.tableStyleInfo = style_resumen
            ws_resumen.add_table(table_resumen)
            
            # Formato de moneda para la columna de total
            for row in ws_resumen.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = money_format
            
            # Formato de porcentaje
            for row in ws_resumen.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '0.00%'
            
            # Añadir fila de total
            total_row = len(total_por_tipo) + 2
            ws_resumen.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            
            cell_total = ws_resumen.cell(row=total_row, column=2, value=float(total_general))
            cell_total.number_format = money_format
            cell_total.font = Font(bold=True)
            
            cell_porcentaje = ws_resumen.cell(row=total_row, column=3, value=1.0 if total_general > 0 else 0)
            cell_porcentaje.number_format = '0.00%'
            cell_porcentaje.font = Font(bold=True)
            
            # Ajustar ancho de columnas
            for column in ws_resumen.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        # Considerar el ancho del encabezado también
                        value_length = len(str(cell.value).encode('utf-8'))  # Contar bytes para caracteres especiales
                        if value_length > max_length:
                            max_length = value_length
                    except:
                        pass
                # Añadir un poco más de espacio para el borde y padding
                adjusted_width = (max_length + 4) * 1.1
                ws_resumen.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Ajustado el ancho máximo a 50
            
            # 5. Hoja de Resumen Mensual
            ws_mensual = wb.create_sheet("Resumen Mensual")
            
            # Calcular totales por mes y año
            total_por_mes = defaultdict(float)
            for factura in self.facturas:
                try:
                    fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    mes_anio = fecha.strftime('%Y-%m')
                    total_por_mes[mes_anio] += factura['valor']
                except Exception as e:
                    logger.warning(f"Error al procesar fecha: {factura['fecha']} - {str(e)}")
                    continue
            
            # Ordenar por fecha
            total_por_mes = dict(sorted(total_por_mes.items()))
            
            # Escribir encabezados
            ws_mensual.append(["Mes", "Año", "Total (COP)"])
            
            # Escribir datos
            for mes_anio, total in total_por_mes.items():
                anio, mes = mes_anio.split('-')
                nombre_mes = [
                    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                ][int(mes) - 1]
                ws_mensual.append([nombre_mes, anio, total])
            
            # Aplicar formato a la tabla de resumen mensual
            if len(total_por_mes) > 0:
                table_mensual = Table(displayName="TablaResumenMensual", ref=f"A1:C{len(total_por_mes) + 1}")
                style_mensual = TableStyleInfo(name="TableStyleMedium3", showFirstColumn=False,
                                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table_mensual.tableStyleInfo = style_mensual
                ws_mensual.add_table(table_mensual)
                
                # Formato de moneda para la columna de total
                for row in ws_mensual.iter_rows(min_row=2, min_col=3, max_col=3):
                    for cell in row:
                        cell.number_format = money_format
                
                # Ajustar ancho de columnas
                for column in ws_mensual.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            # Considerar el ancho del encabezado también
                            value_length = len(str(cell.value).encode('utf-8'))  # Contar bytes para caracteres especiales
                            if value_length > max_length:
                                max_length = value_length
                        except:
                            pass
                    # Añadir un poco más de espacio para el borde y padding
                    adjusted_width = (max_length + 4) * 1.1
                    ws_mensual.column_dimensions[column_letter].width = min(adjusted_width, 40)  # Ajustado el ancho máximo
            
            # Guardar el archivo con configuración para evitar advertencias
            temp_file = file_path + '.tmp'
            try:
                # Guardar primero en un archivo temporal
                wb.save(temp_file)
                
                # Si el archivo de destino existe, eliminarlo
                if os.path.exists(file_path):
                    os.remove(file_path)
                
                # Renombrar el archivo temporal al nombre final
                os.rename(temp_file, file_path)
                
                # Actualizar el directorio de exportación en la configuración
                config['APP']['last_export_dir'] = str(Path(file_path).parent)
                save_config(config)
                
                # Mostrar mensaje de éxito
                QMessageBox.information(
                    self,
                    "Exportación exitosa",
                    f"Los datos se han exportado correctamente a:\n{file_path}"
                )
                
                logger.info(f"Datos exportados exitosamente a {file_path}")
                
            except Exception as e:
                # Si hay un error, intentar limpiar el archivo temporal
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except:
                        pass
                raise e
            
        except PermissionError:
            error_msg = "No se pudo guardar el archivo. Asegúrese de que el archivo no esté abierto en otro programa."
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Error de permisos", error_msg)
            
        except Exception as e:
            error_msg = f"Ocurrió un error al exportar a Excel: {str(e)}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(
                self,
                "Error al exportar",
                f"{error_msg}\n\nPor favor, revise los logs para más detalles."
            )
            # Encabezados
            headers = ["Tipo de Gasto", "Total (COP)", "Porcentaje"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_resumen.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos
            total_general = sum(total_por_tipo.values())
            for row_num, (tipo, total) in enumerate(total_por_tipo.items(), 2):
                ws_resumen.cell(row=row_num, column=1, value=tipo)
                
                # Formato de moneda para el total
                cell_total = ws_resumen.cell(row=row_num, column=2, value=float(total))
                cell_total.number_format = money_format
                
                # Porcentaje
                if total_general > 0:
                    porcentaje = (total / total_general) * 100
                    cell_porcentaje = ws_resumen.cell(row=row_num, column=3, value=porcentaje/100)
                    cell_porcentaje.number_format = '0.00%'
            
            # Fila de total
            row_num = len(total_por_tipo) + 2
            ws_resumen.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
            
            cell_total = ws_resumen.cell(row=row_num, column=2, value=float(total_general))
            cell_total.number_format = money_format
            cell_total.font = Font(bold=True)
            
            # Ajustar ancho de columnas
            for col in ws_resumen.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                ws_resumen.column_dimensions[column].width = min(adjusted_width, 30)
            
            # 3. Hoja de Resumen Mensual
            ws_mensual = wb.create_sheet("Resumen Mensual")
            
            # Calcular totales por mes y tipo
            total_por_mes_tipo = defaultdict(lambda: defaultdict(float))
            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            
            for factura in self.facturas:
                try:
                    fecha = datetime.strptime(factura['fecha'], '%d/%m/%Y')
                    mes = fecha.month - 1  # 0-11
                    total_por_mes_tipo[mes][factura['tipo']] += factura['valor']
                except:
                    continue
            
            # Obtener todos los tipos únicos
            tipos = sorted({tipo for factura in self.facturas for tipo in [factura['tipo']]})
            
            # Encabezados
            headers = ["Mes"] + tipos + ["Total"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_mensual.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Datos por mes
            for mes in range(12):
                if mes in total_por_mes_tipo:
                    row_num = mes + 2  # +2 porque la fila 1 es el encabezado
                    ws_mensual.cell(row=row_num, column=1, value=meses[mes])
                    
                    # Valores por tipo
                    total_mes = 0
                    for col_num, tipo in enumerate(tipos, 2):  # Empieza en columna 2
                        valor = total_por_mes_tipo[mes].get(tipo, 0)
                        if valor > 0:
                            cell = ws_mensual.cell(row=row_num, column=col_num, value=float(valor))
                            cell.number_format = money_format
                            total_mes += valor
                    
                    # Total del mes
                    cell_total = ws_mensual.cell(row=row_num, column=len(tipos)+2, value=float(total_mes))
                    cell_total.number_format = money_format
            
            # Fila de totales por tipo
            row_num = 14  # Después de los 12 meses
            ws_mensual.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
            
            # Calcular totales por tipo
            for col_num, tipo in enumerate(tipos, 2):
                total_tipo = sum(total_por_mes_tipo[mes].get(tipo, 0) for mes in range(12))
                if total_tipo > 0:
                    cell = ws_mensual.cell(row=row_num, column=col_num, value=float(total_tipo))
                    cell.number_format = money_format
                    cell.font = Font(bold=True)
            
            # Total general
            cell_total = ws_mensual.cell(row=row_num, column=len(tipos)+2, 
                                       value=float(sum(sum(m.values()) for m in total_por_mes_tipo.values())))
            cell_total.number_format = money_format
            cell_total.font = Font(bold=True)
            
            # Ajustar ancho de columnas
            for col in ws_mensual.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                ws_mensual.column_dimensions[column].width = min(adjusted_width, 20)
            
            # Guardar el archivo
            wb.save(file_path)
            
            # Mostrar mensaje de éxito
            QMessageBox.information(
                self, 
                "Exportación exitosa", 
                f"Los datos se han exportado correctamente a:\n{file_path}"
            )
            
            logger.info(f"Datos exportados exitosamente a {file_path}")
            
        except Exception as e:
            error_msg = f"Error al exportar a Excel: {str(e)}"
    
    def actualizar_boton_eliminar(self):
        """Actualizar el estado del botón de eliminar basado en la selección"""
        try:
            # Verificar si el botón existe
            if not hasattr(self, 'btn_eliminar'):
                return
                
            # Obtener las filas seleccionadas
            selected_items = self.tabla_facturas.selectedItems()
            has_selection = len(selected_items) > 0
            
            # Actualizar el estado del botón
            self.btn_eliminar.setEnabled(has_selection)
            
            # Debug
            print(f"Botón de eliminar: {'Habilitado' if has_selection else 'Deshabilitado'}")
            if has_selection:
                rows = set(item.row() for item in selected_items)
                print(f"Filas seleccionadas: {rows}")
                
        except Exception as e:
            print(f"Error en actualizar_boton_eliminar: {str(e)}")
            if hasattr(self, 'btn_eliminar'):
                self.btn_eliminar.setEnabled(False)
    
    def actualizar_otra_tabla(self, factura_id, campo, nuevo_valor, es_tabla_filtrada):
        """
        Actualiza el valor en la otra tabla cuando se realiza un cambio en una de ellas
        
        Args:
            factura_id: ID de la factura que se está actualizando
            campo: Nombre del campo que se modificó
            nuevo_valor: Nuevo valor del campo
            es_tabla_filtrada: Booleano que indica si el cambio vino de la tabla filtrada
        """
        try:
            # Determinar qué tabla actualizar (la que no generó el cambio)
            tabla_destino = self.tabla_filtrada if not es_tabla_filtrada else self.tabla_facturas
            
            # Mapear campos a columnas según la tabla de destino
            # La tabla principal tiene columnas: [ID, Fecha, Tipo, Descripción, Valor]
            # La tabla filtrada tiene columnas: [Fecha, Tipo, Descripción, Valor]
            if es_tabla_filtrada:
                # Si el destino es la tabla filtrada (origen: tabla principal)
                mapeo_campos = {
                    'fecha': 0,   # Columna 0 en tabla filtrada
                    'tipo': 1,    # Columna 1 en tabla filtrada
                    'descripcion': 2,  # Columna 2 en tabla filtrada
                    'valor': 3     # Columna 3 en tabla filtrada
                }
            else:
                # Si el destino es la tabla principal (origen: tabla filtrada)
                mapeo_campos = {
                    'fecha': 1,   # Columna 1 en tabla principal (0 es ID)
                    'tipo': 2,    # Columna 2 en tabla principal
                    'descripcion': 3,  # Columna 3 en tabla principal
                    'valor': 4     # Columna 4 en tabla principal
                }
            
            # Obtener el índice de la columna
            columna = mapeo_campos.get(campo)
            if columna is None:
                return
                
            # Buscar la fila que contiene la factura en la tabla de destino
            for fila in range(tabla_destino.rowCount()):
                # Obtener el ID de la factura de la fila actual
                item_id = tabla_destino.item(fila, 0)
                if not item_id:
                    continue
                    
                # Obtener el ID de la factura, ya sea de UserRole o del texto
                if es_tabla_filtrada:
                    # Para la tabla filtrada, el ID está en UserRole
                    current_id = item_id.data(Qt.ItemDataRole.UserRole)
                else:
                    # Para la tabla principal, el ID está en el texto de la columna 0
                    current_id = int(item_id.text()) if item_id.text().isdigit() else None
                
                if current_id == factura_id:
                    # Actualizar la celda correspondiente
                    item = tabla_destino.item(fila, columna)
                    if item:
                        # Bloquear señales temporalmente para evitar bucles
                        tabla_destino.blockSignals(True)
                        try:
                            if campo == 'valor':
                                # Formatear valor numérico
                                try:
                                    valor = float(nuevo_valor)
                                    item.setText(f"${valor:,.2f} COP".replace(",", "X").replace(".", ",").replace("X", "."))
                                except (ValueError, TypeError):
                                    item.setText(str(nuevo_valor))
                            else:
                                item.setText(str(nuevo_valor))
                        finally:
                            # Restaurar señales
                            tabla_destino.blockSignals(False)
                        # Salir del bucle una vez actualizado
                        break
        except Exception as e:
            logger.error(f"Error al actualizar la otra tabla: {str(e)}")

    def guardar_cambios_celda(self, item):
        """
        Maneja los cambios en las celdas editables de las tablas.
        
        Args:
            item: El ítem de la tabla que fue modificado
        """
        # Evitar llamadas recursivas durante la actualización
        if not hasattr(self, '_updating_cell'):
            self._updating_cell = False
            
        if self._updating_cell:
            return
            
        try:
            self._updating_cell = True
            
            # Determinar qué tabla generó el evento
            tabla = self.sender()
            if tabla not in [self.tabla_facturas, self.tabla_filtrada]:
                return
                
            es_tabla_filtrada = (tabla == self.tabla_filtrada)
            
            # Mapeo de columnas a campos de factura según la tabla
            # (es_tabla_filtrada, column_index): 'field_name'
            column_mapping = {
                # Tabla filtrada (sin columna ID)
                (True, 0): 'fecha',      # Columna 0: Fecha
                (True, 1): 'tipo',       # Columna 1: Tipo
                (True, 2): 'descripcion', # Columna 2: Descripción
                (True, 3): 'valor',      # Columna 3: Valor
                # Tabla principal (con columna ID)
                (False, 0): 'id',        # Columna 0: ID (no editable)
                (False, 1): 'fecha',     # Columna 1: Fecha
                (False, 2): 'tipo',      # Columna 2: Tipo
                (False, 3): 'descripcion', # Columna 3: Descripción
                (False, 4): 'valor'      # Columna 4: Valor
            }
            
            # Obtener el campo que se está editando
            campo = column_mapping.get((es_tabla_filtrada, item.column()))
            if not campo or campo == 'id':  # No permitir editar el ID
                return
                
            # Obtener el nuevo valor y limpiarlo
            nuevo_valor = item.text().strip()
            
            # Validar que el campo no esté vacío
            if not nuevo_valor:
                QMessageBox.warning(self, "Error", "El campo no puede estar vacío")
                self._restaurar_valor_anterior(tabla, item, campo)
                return
                
            # Validaciones específicas por tipo de campo
            if campo == 'fecha':
                try:
                    # Validar formato de fecha (DD/MM/YYYY)
                    datetime.strptime(nuevo_valor, '%d/%m/%Y')
                except ValueError:
                    QMessageBox.warning(self, "Formato inválido", 
                                     "El formato de fecha debe ser DD/MM/YYYY")
                    self._restaurar_valor_anterior(tabla, item, campo)
                    return
                    
            elif campo == 'valor':
                try:
                    # Extraer solo los números y el punto decimal
                    valor_limpio = ''.join(c for c in nuevo_valor if c.isdigit() or c in '.,')
                    # Reemplazar comas por puntos para el float
                    valor_limpio = valor_limpio.replace(',', '.')
                    valor_numerico = float(valor_limpio)
                    if valor_numerico <= 0:
                        raise ValueError("El valor debe ser mayor a cero")
                    
                    # Formatear el valor para mostrarlo en la tabla
                    valor_formateado = f"${valor_numerico:,.0f} COP".replace(',', '.')
                    item.setText(valor_formateado)
                    nuevo_valor = str(valor_numerico)  # Guardar como string numérico
                except (ValueError, TypeError) as e:
                    QMessageBox.warning(self, "Valor inválido", 
                                     f"El valor debe ser un número mayor a cero: {str(e)}")
                    self._restaurar_valor_anterior(tabla, item, campo)
                    return
            
            # Obtener el ID de la factura usando el método _obtener_id_factura
            fila = item.row()
            es_tabla_filtrada = (tabla == self.tabla_filtrada) if hasattr(self, 'tabla_filtrada') else False
            factura_id = self._obtener_id_factura(tabla, fila, es_tabla_filtrada)
            
            if factura_id is None:
                logger.error("No se pudo obtener el ID de la factura")
                self._restaurar_valor_anterior(tabla, item, campo)
                return
            
            # Encontrar la factura en la lista de facturas
            factura = next((f for f in self.facturas if f.get('id') == factura_id), None)
            if not factura:
                logger.error(f"No se encontró la factura con ID {factura_id}")
                self._restaurar_valor_anterior(tabla, item, campo)
                return
            
            # Verificar si el valor realmente cambió
            if campo in factura and str(factura[campo]) == nuevo_valor:
                return
            
            # Guardar el valor anterior para restaurar en caso de error
            valor_anterior = factura.get(campo)
            
            try:
                # Actualizar el valor en el diccionario de la factura
                if campo == 'valor':
                    # Para el campo valor, guardar como float
                    factura[campo] = float(nuevo_valor) if isinstance(nuevo_valor, (int, float)) else float(nuevo_valor.replace('.', '').replace(',', '.'))
                else:
                    # Para los demás campos, guardar como string
                    factura[campo] = str(nuevo_valor).strip()
                
                # Guardar los cambios en la base de datos
                if self.db.actualizar_factura(
                    factura_id=factura_id,
                    fecha=factura.get('fecha', ''),
                    tipo=factura.get('tipo', ''),
                    descripcion=factura.get('descripcion', ''),
                    valor=float(factura.get('valor', 0))
                ):
                    # Actualizar la otra tabla si es necesario
                    self.actualizar_otra_tabla(factura_id, campo, nuevo_valor, es_tabla_filtrada)
                    
                    # Actualizar la lista de facturas para reflejar los cambios
                    self.actualizar_lista_facturas()
                    
                    # Actualizar resúmenes
                    self.actualizar_resumen()
                    
                    # Mostrar mensaje de éxito en la barra de estado
                    self.statusBar().showMessage("Cambios guardados correctamente", 3000)
                    
                    # Forzar actualización visual de las tablas
                    tabla.viewport().update()
                    if es_tabla_filtrada:
                        self.tabla_facturas.viewport().update()
                    else:
                        self.tabla_filtrada.viewport().update()
                else:
                    raise Exception("No se pudo actualizar la base de datos")
                    
            except Exception as e:
                # Revertir el cambio en la interfaz
                factura[campo] = valor_anterior
                self._restaurar_valor_anterior(tabla, item, campo, valor_anterior)
                
                # Mostrar mensaje de error
                error_msg = f"Error al guardar los cambios: {str(e)}"
                logger.error(error_msg, exc_info=True)
                QMessageBox.critical(self, "Error", error_msg)
                
        finally:
            self._updating_cell = False
    
    def _restaurar_valor_anterior(self, tabla, item, campo, valor_anterior=None):
        """
        Restaura el valor anterior de una celda después de un error de validación
        
        Args:
            tabla: Referencia a la tabla (self.tabla_facturas o self.tabla_filtrada)
            item: El ítem de la tabla que se está editando
            campo: Nombre del campo que se está editando
            valor_anterior: Valor anterior a restaurar (opcional)
        """
        try:
            tabla.blockSignals(True)
            
            if campo == 'valor' and valor_anterior is not None:
                # Formatear valor numérico
                try:
                    valor = float(valor_anterior)
                    item.setText(f"${valor:,.0f} COP".replace(',', '.'))
                except (ValueError, TypeError):
                    item.setText(str(valor_anterior))
            elif valor_anterior is not None:
                item.setText(str(valor_anterior))
        except Exception as e:
            logger.error(f"Error al restaurar valor anterior: {str(e)}")
        finally:
            tabla.blockSignals(False)

    def _procesar_valor(self, nuevo_valor, factura):
        """
        Procesa el valor numérico de una factura, validando y formateando correctamente.
        
        Args:
            nuevo_valor: El valor a procesar (puede ser string con formato de moneda)
            factura: Diccionario con los datos de la factura
            
        Returns:
            float: El valor numérico procesado
        """
        try:
            # Limpiar el valor numérico
            valor_limpio = str(nuevo_valor).replace('$', '').replace('COP', '').replace('.', '').replace(' ', '').replace(',', '.')
            valor_numerico = round(float(valor_limpio), 2)
            if valor_numerico < 0:
                raise ValueError("El valor no puede ser negativo")
            return valor_numerico
        except (ValueError, TypeError) as e:
            # Si hay un error, devolver el valor anterior de la factura
            return float(factura.get('valor', 0))

    def _formatear_valor_moneda(self, valor):
        """
        Formatea un valor numérico como moneda colombiana.
        
        Args:
            valor: Valor numérico a formatear
            
        Returns:
            str: Valor formateado como moneda colombiana
        """
        try:
            valor_float = float(valor)
            return f"${valor_float:,.0f} COP".replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            return "$0 COP"

    def _obtener_id_factura(self, tabla, fila, es_tabla_filtrada):
        """
        Obtiene el ID de la factura desde una fila de la tabla.
        
        Args:
            tabla: Tabla de la que obtener el ID
            fila: Índice de la fila
            es_tabla_filtrada: Indica si es la tabla filtrada
            
        Returns:
            int or None: ID de la factura o None si no se pudo obtener
        """
        try:
            # Obtener el ítem de la primera columna (donde normalmente está el ID)
            item = tabla.item(fila, 0)
            if item is None:
                logger.warning("No se pudo obtener el ítem de la tabla")
                return None
                
            # Primero intentar obtener el ID del UserRole (usado en la tabla filtrada)
            factura_id = item.data(Qt.ItemDataRole.UserRole)
            if factura_id is not None and str(factura_id).isdigit():
                return int(factura_id)
                
            # Si no está en UserRole, intentar obtenerlo del texto
            if item.text().isdigit():
                return int(item.text())
                
            # Si no se pudo obtener, intentar con la columna oculta (si existe)
            if tabla.columnCount() > 4:  # Si hay una columna oculta con el ID
                item_id = tabla.item(fila, 4)  # Asumiendo que el ID está en la columna 4 (oculta)
                if item_id and item_id.text().isdigit():
                    return int(item_id.text())
                    
            logger.warning(f"No se pudo obtener el ID de la factura de la fila {fila}")
            return None
            
        except Exception as e:
            logger.error(f"Error al obtener ID de factura: {str(e)}", exc_info=True)
            return None
        
        try:
            # Actualizar el valor en el diccionario de la factura
            factura[campo] = nuevo_valor
            
            # Guardar los cambios en la base de datos
            if self.db.actualizar_factura(
                factura_id=factura_id,
                fecha=factura['fecha'],
                tipo=factura['tipo'],
                descripcion=factura['descripcion'],
                valor=factura['valor']
            ):
                # Mostrar mensaje de éxito
                self.statusBar().showMessage("Cambios guardados correctamente", 2000)
                
                # Resaltar la fila modificada temporalmente
                for col in range(tabla.columnCount()):
                    cell_item = tabla.item(row, col)
                    if cell_item:
                        cell_item.setBackground(QColor(230, 255, 230))  # Verde claro
                
                # Programar la restauración del color después de 2 segundos
                QTimer.singleShot(2000, lambda r=row, t=tabla: self.restaurar_color_fila(r, 0, t))
                
                # Actualizar la otra tabla si es necesario
                self.actualizar_otra_tabla(factura_id, campo, nuevo_valor, es_tabla_filtrada)
                
            else:
                # Si hay un error al guardar, restaurar el valor original
                raise Exception("No se pudo guardar el cambio en la base de datos")
                
        except Exception as e:
            # Restaurar el valor anterior en caso de error
            if campo in factura and valor_anterior is not None:
                factura[campo] = valor_anterior
                if campo == 'valor':
                    item.setText(f"${float(valor_anterior):,.2f} COP".replace(',', 'X').replace('.', ',').replace('X', '.'))
                else:
                    item.setText(str(valor_anterior))
            
            QMessageBox.critical(self, "Error", str(e))
                
        finally:
            # Restaurar las señales de la tabla
            tabla.blockSignals(False)
    
    def restaurar_color_fila(self, row, col, tabla):
        """
        Restaurar el color de fondo de una fila después de una modificación
        
        Args:
            row: Índice de la fila
            col: Índice de la columna (no utilizado, se mantiene por compatibilidad)
            tabla: Referencia a la tabla (self.tabla_facturas o self.tabla_filtrada)
        """
        if not tabla or row < 0 or row >= tabla.rowCount():
            return
            
        # Obtener el ID de la factura de la fila
        id_item = tabla.item(row, 0)
        if not id_item:
            return
            
        factura_id = None
        if tabla == self.tabla_facturas:
            try:
                factura_id = int(id_item.text())
            except (ValueError, AttributeError):
                return
        else:  # tabla_filtrada
            factura_id = id_item.data(Qt.ItemDataRole.UserRole)
            if not factura_id:
                return
        
        # Buscar la factura correspondiente
        factura = next((f for f in self.facturas if f.get('id') == factura_id), None)
        if not factura:
            return
        
        # Restaurar el color de fondo según el valor
        color_fondo = QColor(255, 255, 255)  # Blanco por defecto
        if factura.get('valor', 0) > 1000000:  # Resaltar facturas mayores a 1 millón
            color_fondo = QColor(255, 230, 230)  # Rojo claro
        
        # Aplicar el color a todas las celdas de la fila
        for col in range(tabla.columnCount()):
            cell_item = tabla.item(row, col)
            if cell_item:
                cell_item.setBackground(color_fondo)
    
    def eliminar_facturas_seleccionadas(self):
        """Eliminar las facturas seleccionadas de la lista"""
        print("Botón 'Eliminar seleccionadas' presionado")  # Debug
        
        # Obtener las filas seleccionadas (sin duplicados)
        selected_ranges = self.tabla_facturas.selectedRanges()
        filas_seleccionadas = set()
        
        print(f"Rangos seleccionados: {selected_ranges}")  # Debug
        
        for range_ in selected_ranges:
            filas_seleccionadas.update(range(range_.topRow(), range_.bottomRow() + 1))
        
        print(f"Filas seleccionadas: {filas_seleccionadas}")  # Debug
        
        if not filas_seleccionadas:
            print("No hay filas seleccionadas")  # Debug
            QMessageBox.warning(self, "Eliminar Facturas", "No hay filas seleccionadas para eliminar.")
            return
        
        # Obtener los IDs de las facturas seleccionadas
        facturas_a_eliminar = []
        for fila in filas_seleccionadas:
            if 0 <= fila < self.tabla_facturas.rowCount():
                id_item = self.tabla_facturas.item(fila, 0)  # ID está en la columna 0 (oculta)
                print(f"Fila {fila}: id_item = {id_item}")  # Debug
                if id_item is not None:
                    try:
                        factura_id = int(id_item.text())
                        facturas_a_eliminar.append(factura_id)
                        print(f"ID de factura encontrado: {factura_id}")  # Debug
                    except (ValueError, AttributeError) as e:
                        print(f"Error al obtener ID de factura: {e}")  # Debug
                        continue
        
        print(f"Facturas a eliminar: {facturas_a_eliminar}")  # Debug
        
        if not facturas_a_eliminar:
            print("No se encontraron IDs de facturas válidas")  # Debug
            QMessageBox.warning(self, "Eliminar Facturas", "No se pudieron identificar las facturas a eliminar.")
            return
        
        # Confirmar eliminación
        confirmacion = QMessageBox.question(
            self,
            "Confirmar Eliminación",
            f"¿Está seguro de que desea eliminar {len(facturas_a_eliminar)} factura(s) seleccionada(s)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if confirmacion == QMessageBox.StandardButton.Yes:
            print(f"Confirmada eliminación de {len(facturas_a_eliminar)} facturas")  # Debug
            # Eliminar las facturas de la base de datos
            eliminaciones_exitosas = 0
            try:
                for factura_id in facturas_a_eliminar:
                    print(f"Intentando eliminar factura con ID: {factura_id}")  # Debug
                    if self.db.eliminar_factura(factura_id):
                        eliminaciones_exitosas += 1
                        print(f"Factura {factura_id} eliminada correctamente")  # Debug
                    else:
                        print(f"No se pudo eliminar la factura {factura_id}")  # Debug
                
                # Actualizar la interfaz
                print("Recargando datos...")  # Debug
                self.cargar_datos(actualizar_ui=True)
                
                mensaje = f"Se eliminaron {eliminaciones_exitosas} de {len(facturas_a_eliminar)} factura(s) correctamente."
                self.statusBar().showMessage(mensaje, 5000)  # 5 segundos
                print(mensaje)  # Debug
                logger.info(mensaje)
                
                # Mostrar mensaje si no se pudieron eliminar todas las facturas
                if eliminaciones_exitosas < len(facturas_a_eliminar):
                    QMessageBox.warning(
                        self,
                        "Advertencia",
                        f"Solo se pudieron eliminar {eliminaciones_exitosas} de {len(facturas_a_eliminar)} facturas seleccionadas."
                    )
                
            except Exception as e:
                error_msg = f"Error al eliminar las facturas: {str(e)}"
                print(error_msg)  # Debug
                logger.error(error_msg, exc_info=True)
                QMessageBox.critical(self, "Error", error_msg)
    
    def confirmar_limpiar_todo(self):
        """Mostrar diálogo de confirmación para limpiar todos los datos"""
        if not self.facturas:
            QMessageBox.information(self, "Limpiar Datos", "No hay datos para limpiar.")
            return
        
        confirmacion = QMessageBox.warning(
            self,
            "Confirmar Limpieza Total",
            "¿Está seguro de que desea eliminar TODAS las facturas?\n\n"
            "Esta acción NO se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if confirmacion == QMessageBox.StandardButton.Yes:
            self.limpiar_todo()
    
    def limpiar_todo(self):
        """Eliminar todas las facturas"""
        # Guardar una copia de respaldo
        facturas_backup = self.facturas.copy()
        
        try:
            # Limpiar la lista de facturas
            self.facturas.clear()
            
            # Guardar los cambios
            if self.guardar_datos():
                # Actualizar la interfaz
                self.actualizar_lista_facturas()
                self.actualizar_resumen()
                self.statusBar().showMessage("Se eliminaron todas las facturas correctamente.", 3000)
                logger.info("Se eliminaron todas las facturas")
        except Exception as e:
            # Restaurar la copia de respaldo en caso de error
            self.facturas = facturas_backup
            error_msg = f"Error al limpiar los datos: {str(e)}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Error", error_msg)
    
    def importar_desde_json(self):
        """Importar facturas desde un archivo JSON"""
        # Abrir diálogo para seleccionar archivo
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo JSON",
            "",
            "Archivos JSON (*.json);;Todos los archivos (*)"
        )
        
        if not file_path:
            return  # Usuario canceló el diálogo
        
        try:
            # Leer el archivo JSON
            with open(file_path, 'r', encoding='utf-8') as f:
                facturas_importadas = json.load(f)
            
            if not isinstance(facturas_importadas, list):
                raise ValueError("El archivo no contiene una lista de facturas válida.")
            
            # Validar el formato de las facturas
            facturas_validas = []
            for i, factura in enumerate(facturas_importadas, 1):
                if not all(key in factura for key in ['fecha', 'tipo', 'descripcion', 'valor']):
                    logger.warning(f"Factura {i} omitida: formato incorrecto")
                    continue
                
                try:
                    # Validar y convertir los valores
                    fecha_valida = bool(datetime.strptime(factura['fecha'], '%d/%m/%Y'))
                    valor_valido = float(factura['valor']) > 0
                    
                    if fecha_valida and valor_valido:
                        # Asegurar que el valor sea un número
                        factura['valor'] = float(factura['valor'])
                        facturas_validas.append(factura)
                    else:
                        logger.warning(f"Factura {i} omitida: fecha o valor inválidos")
                except (ValueError, TypeError):
                    logger.warning(f"Factura {i} omitida: error en los datos")
            
            if not facturas_validas:
                QMessageBox.warning(self, "Importar Datos", "No se encontraron facturas válidas para importar.")
                return
            
            # Mostrar resumen de importación
            resumen = f"Se importarán {len(facturas_validas)} factura(s) de {len(facturas_importadas)} encontradas."
            if len(facturas_validas) < len(facturas_importadas):
                resumen += "\n\nAlgunas facturas se omitieron por tener formato incorrecto."
            
            confirmacion = QMessageBox.question(
                self,
                "Confirmar Importación",
                f"{resumen}\n\n¿Desea continuar con la importación?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if confirmacion == QMessageBox.StandardButton.Yes:
                # Hacer una copia de respaldo
                facturas_originales = self.facturas.copy()
                
                try:
                    # Agregar las facturas importadas a la lista existente
                    self.facturas.extend(facturas_validas)
                    
                    # Guardar los cambios
                    if self.guardar_datos():
                        # Actualizar la interfaz
                        self.actualizar_lista_facturas()
                        self.actualizar_resumen()
                        self.actualizar_filtros()
                        
                        # Mostrar mensaje de éxito
                        QMessageBox.information(
                            self,
                            "Importación Exitosa",
                            f"Se importaron {len(facturas_validas)} factura(s) correctamente."
                        )
                        
                        logger.info(f"Se importaron {len(facturas_validas)} facturas desde {file_path}")
                        self.statusBar().showMessage("Importación completada correctamente.", 3000)
                except Exception as e:
                    # Restaurar la copia de respaldo en caso de error
                    self.facturas = facturas_originales
                    error_msg = f"Error al importar las facturas: {str(e)}"
                    logger.error(error_msg, exc_info=True)
                    QMessageBox.critical(self, "Error", error_msg)
        
        except json.JSONDecodeError:
            QMessageBox.critical(self, "Error", "El archivo seleccionado no es un JSON válido.")
        except Exception as e:
            error_msg = f"Error al importar desde JSON: {str(e)}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(self, "Error", error_msg)
    
    def cargar_preferencia_tema(self):
        """Cargar la preferencia de tema desde el archivo de configuración"""
        config_path = self.data_dir / "tema_config.json"
        try:
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    if 'tema_oscuro' in config:
                        self.tema_oscuro = config['tema_oscuro']
                        self.aplicar_tema()
        except Exception as e:
            logger.warning(f"No se pudo cargar la preferencia de tema: {str(e)}")
    
    def guardar_preferencia_tema(self):
        """Guardar la preferencia de tema en el archivo de configuración"""
        config_path = self.data_dir / "tema_config.json"
        try:
            config = {}
            if config_path.exists():
                try:
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                except (json.JSONDecodeError, Exception) as e:
                    logger.warning(f"Error al leer el archivo de configuración del tema: {str(e)}")
                    config = {}
            
            config['tema_oscuro'] = self.tema_oscuro
            
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=4)
        except Exception as e:
            logger.error(f"Error al guardar la preferencia de tema: {str(e)}")
    
    def cambiar_tema(self):
        """Alternar entre modo oscuro y claro"""
        self.tema_oscuro = not self.tema_oscuro
        # Actualizar el texto del botón según el tema
        self.btn_tema.setText("Modo Claro" if self.tema_oscuro else "Modo Oscuro")
        self.aplicar_tema()
        self.guardar_preferencia_tema()
    
    def aplicar_tema(self):
        """Aplicar el tema seleccionado a toda la aplicación con diseño moderno"""
        # Establecer ancho mínimo para los botones de acción
        if hasattr(self, 'menu_importar') and hasattr(self, 'btn_limpiar_todo') and hasattr(self, 'btn_exportar'):
            for btn in [self.menu_importar, self.btn_limpiar_todo, self.btn_exportar]:
                btn.setMinimumWidth(150)  # Ancho mínimo de 150 píxeles
                btn.setSizePolicy(QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Preferred)
                
        if self.tema_oscuro:
            # Estilo moderno para modo oscuro
            self.setStyleSheet("""
                /* Estilos generales */
                QMainWindow, QDialog, QWidget {
                    background-color: #1a1a2e;
                    color: #e6e6e6;
                    font-family: 'Segoe UI', Arial, sans-serif;
                }
                
                /* Estilos para menús desplegables - Modo Oscuro */
                QMenu {
                    background-color: #2d3748;  /* Fondo oscuro para el menú */
                    color: #e2e8f0;  /* Texto claro */
                    border: 1px solid #4a5568;  /* Borde más oscuro */
                    padding: 8px;
                    border-radius: 4px;
                    min-width: 200px;
                }
                
                QMenu::item {
                    padding: 8px 25px 8px 20px;
                    border: 1px solid #4a5568;  /* Borde para modo oscuro */
                    min-width: 160px;
                    border-radius: 3px;
                    margin: 3px 0;
                    background-color: #2d3748;  /* Fondo oscuro para items */
                }
                
                QMenu::item:selected {
                    background-color: #4a5568;  /* Fondo más claro al seleccionar */
                    color: #ffffff;  /* Texto blanco */
                    border: 1px solid #63b3ed;  /* Borde azul claro */
                    font-weight: 500;
                }
                
                QMenu::item:disabled {
                    color: #718096;  /* Texto gris para deshabilitados */
                }
                
                QMenu::separator {
                    height: 1px;
                    background: #4a5568;  /* Separador más oscuro */
                    margin: 5px 0;
                }
                
                /* Barra de título */
                QLabel[title="true"] {
                    font-size: 24px;
                    font-weight: bold;
                    color: #ffffff;
                    padding: 10px;
                }
            
            /* Contenedor de botones */
            QHBoxLayout {
                spacing: 10px;
            }
            
            /* Botones principales - tamaño uniforme */
            QPushButton#menu_importar,
            QPushButton#btn_limpiar_todo,
            QPushButton#btn_exportar {
                min-width: 120px;
                max-width: 120px;
                padding: 8px 10px;
                margin: 0;
            }
            
            /* Botón Importar Facturas */
            QPushButton#menu_importar {
                background-color: #0f3460;
                color: #ffffff;
            }
            
            /* Botón Exportar a Excel */
            QPushButton#btn_exportar {
                background-color: #0f3460;
                color: #ffffff;
            }
            
            /* Botón Limpiar Todo */
            QPushButton#btn_limpiar_todo {
                background-color: #f8d7da;
                color: #721c24;
                border: 1px solid #f5c6cb;
            }
            
            /* Estados de los botones */
            QPushButton#menu_importar:hover,
            QPushButton#btn_exportar:hover {
                background-color: #1a4b8c;
            }
            
            QPushButton#menu_importar:pressed,
            QPushButton#btn_exportar:pressed {
                background-color: #0d2b4e;
            }
            
            /* Estados específicos para Limpiar Todo */
            QPushButton#btn_limpiar_todo:hover {
                background-color: #f5c6cb;
                border-color: #f1b0b7;
            }
            
            QPushButton#btn_limpiar_todo:pressed {
                background-color: #f1b0b7;
                border-color: #ea99a3;
            }
            
            /* Tarjetas (QGroupBox) */
            QGroupBox {
                background-color: #16213e;
                border: 1px solid #0f3460;
                border-radius: 12px;
                margin-top: 15px;
                padding-top: 25px;
                color: #e6e6e6;
            }
            
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #e94560;
                font-weight: bold;
                font-size: 14px;
            }
            
            /* Botones generales */
            QPushButton {
                background-color: #0f3460;
                color: #ffffff;
                border: none;
                padding: 8px 20px;
                border-radius: 8px;
                font-weight: 500;
                min-width: 100px;
            }
            
            QPushButton:hover {
                background-color: #1a4b8c;
            }
            
            QPushButton:pressed {
                background-color: #0d2b4e;
            }
            
            /* Botón Guardar */
            QPushButton#btn_guardar {
                background-color: #2d4263;
            }
            
            QPushButton#btn_guardar:hover {
                background-color: #1e2f4a;
            }
            
            /* Campos de entrada */
            QLineEdit, QTextEdit, QComboBox, QDateEdit {
                background-color: #16213e;
                color: #e6e6e6;
                border: 1px solid #0f3460;
                padding: 8px;
                border-radius: 6px;
                selection-background-color: #e94560;
                selection-color: #ffffff;
            }
            
            /* Pestañas */
            QTabWidget::pane {
                border: none;
                background: #16213e;
            }
            
            QTabBar::tab {
                background: #1a1a2e;
                color: #a1a1a1;
                padding: 10px 20px;
                border: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 4px;
                font-weight: 500;
            }
            
            QTabBar::tab:selected {
                background: #0f3460;
                color: #ffffff;
            }
            
            QTabBar::tab:!selected {
                margin-top: 4px;
            }
            
            /* Tablas */
            QTableWidget {
                background-color: #16213e;
                color: #e6e6e6;
                gridline-color: #0f3460;
                border: 1px solid #0f3460;
                border-radius: 8px;
                alternate-background-color: #1a1a2e;
            }
            
            QTableWidget::item {
                padding: 8px;
            }
            
            QTableWidget::item:selected {
                background-color: #e94560;
                color: #ffffff;
            }
            
            QHeaderView::section {
                background-color: #0f3460;
                color: #ffffff;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            
            /* Barra de estado */
            QStatusBar {
                background-color: #0f3460;
                color: #ffffff;
                border-top: 1px solid #0d2b4e;
            }
        """)
        else:
            # Estilo moderno para modo claro - versión clara del modo oscuro
            self.setStyleSheet("""
                /* Estilos generales */
                QMainWindow, QDialog, QWidget {
                    background-color: #f0f2f5;
                    color: #2c3e50;
                    font-family: 'Segoe UI', Arial, sans-serif;
                }
                
                /* Estilos para menús desplegables - Modo Claro */
            QMenu {
                background-color: #ffffff;
                color: #2c3e50;
                border: 1px solid #d1d5db;
                padding: 8px;
                border-radius: 4px;
                min-width: 200px;
            }
            
            QMenu::item {
                padding: 8px 25px 8px 20px;
                border: 1px solid #e5e7eb;  /* Borde sutil para modo claro */
                min-width: 160px;
                border-radius: 3px;
                margin: 3px 0;
                background-color: #ffffff;  /* Fondo blanco */
            }
            
            QMenu::item:selected {
                background-color: #f0f9ff;  /* Fondo azul muy claro al seleccionar */
                color: #0369a1;  /* Texto azul oscuro */
                border: 1px solid #7dd3fc;  /* Borde azul claro */
                font-weight: 500;
            }
            
            QMenu::item:disabled {
                color: #9ca3af;
            }
            
            QMenu::separator {
                height: 1px;
                background: #e5e7eb;  /* Separador gris claro */
                margin: 5px 0;
            }
                
                /* Barra de título */
                QLabel[title="true"] {
                    font-size: 24px;
                    font-weight: bold;
                    color: #2c3e50;
                    padding: 10px;
                }
            
            /* Contenedor de botones */
            QHBoxLayout {
                spacing: 10px;
            }
            
            /* Botones principales - tamaño uniforme */
            QPushButton#menu_importar,
            QPushButton#btn_limpiar_todo,
            QPushButton#btn_exportar {
                min-width: 120px;
                max-width: 120px;
                padding: 8px 10px;
                margin: 0;
            }
            
                /* Botón Importar Facturas */
            QPushButton#menu_importar {
                background-color: #3498db;
                color: #ffffff;
                border: 1px solid #2980b9;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: 500;
            }
            
            /* Botón Exportar a Excel */
            QPushButton#btn_exportar {
                background-color: #3498db;
                color: #ffffff;
            }
            
            /* Botón Limpiar Todo */
            QPushButton#btn_limpiar_todo {
                background-color: #f8d7da;
                color: #721c24;
                border: 1px solid #f5c6cb;
            }
            
            /* Estados de los botones */
            QPushButton#menu_importar:hover,
            QPushButton#btn_exportar:hover {
                background-color: #2980b9;
                border-color: #2472a4;
            }
            
            QPushButton#menu_importar:pressed,
            QPushButton#btn_exportar:pressed {
                background-color: #2472a4;
            }
            
            /* Estados específicos para Limpiar Todo */
            QPushButton#btn_limpiar_todo:hover {
                background-color: #f5c6cb;
                border-color: #f1b0b7;
            }
            
            QPushButton#btn_limpiar_todo:pressed {
                background-color: #f1b0b7;
                border-color: #ea99a3;
            }
            
            /* Tarjetas (QGroupBox) */
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #d6dbe2;
                border-radius: 12px;
                margin-top: 15px;
                padding-top: 25px;
                color: #333333;
            }
            
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px;
                color: #3498db;
                font-weight: bold;
                font-size: 14px;
            }
            
            /* Botones generales */
            QPushButton {
                background-color: #3498db;
                color: #ffffff;
                border: none;
                padding: 8px 20px;
                border-radius: 8px;
                font-weight: 500;
                min-width: 100px;
            }
            
            QPushButton:hover {
                background-color: #2980b9;
            }
            
            QPushButton:pressed {
                background-color: #2472a4;
            }
            
            /* Botón Guardar */
            QPushButton#btn_guardar {
                background-color: #2ecc71;
            }
            
            QPushButton#btn_guardar:hover {
                background-color: #27ae60;
            }
            
            /* Campos de entrada */
            QLineEdit, QTextEdit, QComboBox, QDateEdit {
                background-color: #ffffff;
                color: #333333;
                border: 1px solid #d6dbe2;
                padding: 8px;
                border-radius: 6px;
                selection-background-color: #3498db;
                selection-color: #ffffff;
            }
            
            /* Pestañas */
            QTabWidget::pane {
                border: none;
                background: #ffffff;
            }
            
            QTabBar::tab {
                background: #ecf0f1;
                color: #7f8c8d;
                padding: 10px 20px;
                border: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 4px;
                font-weight: 500;
            }
            
            QTabBar::tab:selected {
                background: #3498db;
                color: #ffffff;
            }
            
            QTabBar::tab:!selected {
                margin-top: 4px;
            }
            
            /* Tablas */
            QTableWidget {
                background-color: #ffffff;
                color: #333333;
                gridline-color: #d6dbe2;
                border: 1px solid #d6dbe2;
                border-radius: 8px;
                alternate-background-color: #f8f9fa;
            }
            
            QTableWidget::item {
                padding: 8px;
            }
            
            QTableWidget::item:selected {
                background-color: #3498db;
                color: #ffffff;
            }
            
            QHeaderView::section {
                background-color: #3498db;
                color: #ffffff;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            
            /* Barra de estado */
            QStatusBar {
                background-color: #3498db;
                color: #ffffff;
                border-top: 1px solid #2980b9;
            }
        """)
        
        # Actualizar el texto del botón de tema
        self.btn_tema.setText("Modo Claro" if self.tema_oscuro else "Modo Oscuro")
        
        # Aplicar sombras a los widgets (requiere Qt5)
        self.aplicar_sombras()
        
        # Forzar actualización de la interfaz
        self.update()
    
        # Actualizar todos los widgets hijos
        for widget in self.findChildren(QWidget):
            widget.update()
    
    def aplicar_sombras(self):
        """Aplicar efectos de sombra a los widgets principales"""
        from PyQt6.QtWidgets import QGraphicsDropShadowEffect
        
        # Configurar sombra para tarjetas (GroupBox)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        shadow.setXOffset(0)
        shadow.setYOffset(5)
        shadow.setColor(QColor(0, 0, 0, 80 if self.tema_oscuro else 30))
        
        for group_box in self.findChildren(QGroupBox):
            group_box.setGraphicsEffect(shadow)
        
        # Configurar sombra para la tabla
        table_shadow = QGraphicsDropShadowEffect()
        table_shadow.setBlurRadius(10)
        table_shadow.setXOffset(0)
        table_shadow.setYOffset(3)
        table_shadow.setColor(QColor(0, 0, 0, 50 if self.tema_oscuro else 20))
        
        if hasattr(self, 'tabla_facturas'):
            self.tabla_facturas.setGraphicsEffect(table_shadow)
        
        if hasattr(self, 'tabla_filtrada'):
            self.tabla_filtrada.setGraphicsEffect(table_shadow)

def main():
    app = QApplication(sys.argv)
    
    # Establecer estilo visual
    app.setStyle('Fusion')
    
    # Crear y mostrar ventana principal
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()